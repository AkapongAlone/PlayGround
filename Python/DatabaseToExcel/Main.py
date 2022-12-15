from tkinter.constants import NONE
from typing import Text
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, QObject, pyqtSignal, pyqtSlot, Qt, QPoint,QTimer
from PyQt5.QtGui import QIcon,QMovie
from PyQt5.QtWidgets import (
    QHBoxLayout, QLineEdit, QVBoxLayout, QWidget, QPushButton, QApplication, QGridLayout, QMessageBox, QMainWindow, QFrame,QLabel)
import UI6_Ver as main_verti
import UI6 as main
from select_v import Choosepage_verti
from select_h import Choosepage
from openpyxl import Workbook, load_workbook
import easygui
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import pyodbc
import sys
from tkinter import Tk    
from tkinter.filedialog import askopenfilename
from PyQt5 import uic
import datetime
import subprocess
import os
import yagmail
import re
from datetime import date,timedelta

hwid = str(subprocess.check_output('wmic bios get serialnumber ')).split('\\r\\n')[1].strip('\\r').strip()

print(hwid)

def gen_key(hwid):
  serial=[]
  while len(hwid)<8:
          hwid=hwid+'S'      
  if len(hwid)>8:
          hwid=hwid[:8]   
  for i in hwid:
    if i == 'A':
      serial.append('S')
      serial.append('8')
    elif i =='B':
      serial.append('E')
      serial.append('S')
    elif i =='C':
      serial.append('Q')
      serial.append('D')
    elif i =='D':
      serial.append('1')
      serial.append('L')
    elif i =='E':
      serial.append('8')
      serial.append('W')
    elif i =='F':
      serial.append('P')
      serial.append('K')
    elif i =='G':
      serial.append('J')
      serial.append('3')
    elif i =='H':
      serial.append('Q')
      serial.append('T')
    elif i =='I':
      serial.append('3')
      serial.append('R')
    elif i =='J':
      serial.append('U')
      serial.append('9')
    elif i =='K':
      serial.append('T')
      serial.append('E')
    elif i =='L':
      serial.append('M')
      serial.append('H')
    elif i =='M':
      serial.append('N')
      serial.append('C')
    elif i =='N':
      serial.append('R')
      serial.append('X')
    elif i =='O':
      serial.append('4')
      serial.append('B')
    elif i =='P':
      serial.append('Z')
      serial.append('A')
    elif i =='Q':
      serial.append('V')
      serial.append('M')
    elif i =='R':
      serial.append('Y')
      serial.append('N')
    elif i =='S':
      serial.append('T')
      serial.append('3')
    elif i =='T':
      serial.append('X')
      serial.append('A')
    elif i =='U':
      serial.append('W')
    elif i =='V':
      serial.append('5')
      serial.append('Y')
    elif i =='W':
      serial.append('H')
      serial.append('R')
    elif i =='X':
      serial.append('0')
      serial.append('S')
    elif i =='Y':
      serial.append('G')
      serial.append('0')
    elif i =='Z':
      serial.append('C')
      serial.append('B')
    elif i =='0':
      serial.append('A')
      serial.append('U')
    elif i =='1':
      serial.append('F')
      serial.append('5')
    elif i =='2':
      serial.append('I')
      serial.append('Q')
    elif i =='3':
      serial.append('L')
      serial.append('Y')
    elif i =='4':
      serial.append('D')
      serial.append('V')
    elif i =='5':
      serial.append('2')
      serial.append('S')
    elif i =='6':
      serial.append('6')
      serial.append('R')
    elif i =='7':
      serial.append('9')
      serial.append('H')
    elif i =='8':
      serial.append('7')
      serial.append('B')
    elif i =='9':
      serial.append('E')
      serial.append('J')
  serial_key=serial[0]+serial[1]+serial[2]+serial[3]+"-"+serial[4]+serial[5]+serial[6]+serial[7]+"-"+serial[8]+serial[9]+serial[10]+serial[11]+"-"+serial[12]+serial[13]+serial[14]+serial[15]
  return(serial_key)  
global key_id
key_id=gen_key(hwid)
print(key_id)







class Vertical(QtWidgets.QMainWindow, main_verti.Ui_MainWindow):
    def __init__(self, servername,UID,PWD,db):
        super(Vertical,self).__init__()
        # กำหนดค่าให้mainwindowsให้ใช้ฟังใช่พวกpyqtได้
        # MainWindow = QtWidgets.QMainWindow()
        # self.ui = main_verti.Ui_MainWindow()  # ดึงคลาสจากUI3เข้ามาใช้ในคลาสนี้
        # self.ui.setupUi(MainWindow)  # ผ่านค่าMainwindowsเข้าไปในคลาสsetupUi
        # เรียกใช้ฟังชั่นต่าง ในคลาสsetupUiเข้ามาใช้ในคลาสนี้ได้
        uic.loadUi("UI/Filter_Data.ui", self)
        # self.setupUi(self)
        try : 
            self.conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                                  servername+'; DATABASE='+db+';trusted_connection=yes;' )
            self.cursor = self.conn.cursor()
            
            
        except:
            self.conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; SERVER='+servername+'; DATABASE='+db+';UID='+UID+';PWD='+PWD)
            
            self.cursor = self.conn.cursor()
        Call=[]
        Call.append(self.Cstart)
        Call.append(self.Cstart2)
        Call.append(self.Cstart3)
        Call.append(self.Cstart4)
        Call.append(self.Cstart5)
        Call.append(self.Cstart6)
        Call.append(self.Cstart7)
        Call.append(self.Cstart8)
        Call.append(self.Cstart9)
        Call.append(self.Cstart10)
        Call.append(self.Cstart11)
        Call.append(self.Cstart12)
        Call.append(self.Cstart13)
        Call.append(self.Cstart14)
        Call.append(self.Cstart15)
        Call.append(self.Cstart16)
        Call.append(self.Cstart17)
        Call.append(self.Cstart18)
        Call.append(self.Cstart19)
        Call.append(self.Cstart20)
        Reall=[]
        Reall.append(self.Rend)
        Reall.append(self.Rend2)
        Reall.append(self.Rend3)
        Reall.append(self.Rend4)
        Reall.append(self.Rend5)
        Reall.append(self.Rend6)
        Reall.append(self.Rend7)
        Reall.append(self.Rend8)
        Reall.append(self.Rend9)
        Reall.append(self.Rend10)
        Reall.append(self.Rend11)
        Reall.append(self.Rend12)
        Reall.append(self.Rend13)
        Reall.append(self.Rend14)
        Reall.append(self.Rend15)
        Reall.append(self.Rend16)
        Reall.append(self.Rend17)
        Reall.append(self.Rend18)
        Reall.append(self.Rend19)
        Reall.append(self.Rend20)
        Rall=[]
        Rall.append(self.Rstart)
        Rall.append(self.Rstart2)
        Rall.append(self.Rstart3)
        Rall.append(self.Rstart4)
        Rall.append(self.Rstart5)
        Rall.append(self.Rstart6)
        Rall.append(self.Rstart7)
        Rall.append(self.Rstart8)
        Rall.append(self.Rstart9)
        Rall.append(self.Rstart10)
        Rall.append(self.Rstart11)
        Rall.append(self.Rstart12)
        Rall.append(self.Rstart13)
        Rall.append(self.Rstart14)
        Rall.append(self.Rstart15)
        Rall.append(self.Rstart16)
        Rall.append(self.Rstart17)
        Rall.append(self.Rstart18)
        Rall.append(self.Rstart19)
        Rall.append(self.Rstart20)
        featurall=[]
        featurall.append(self.SheetBox_001)
        featurall.append(self.SheetBox_002)
        featurall.append(self.SheetBox_003)
        featurall.append(self.SheetBox_004)
        featurall.append(self.SheetBox_005)
        featurall.append(self.SheetBox_006)
        featurall.append(self.SheetBox_007)
        featurall.append(self.SheetBox_008)
        featurall.append(self.SheetBox_009)
        featurall.append(self.SheetBox_010)
        featurall.append(self.SheetBox_011)
        featurall.append(self.SheetBox_012)
        featurall.append(self.SheetBox_013)
        featurall.append(self.SheetBox_014)
        featurall.append(self.SheetBox_015)
        featurall.append(self.SheetBox_016)
        featurall.append(self.SheetBox_017)
        featurall.append(self.SheetBox_018)
        featurall.append(self.SheetBox_019)
        featurall.append(self.SheetBox_020)
        self.featurall=featurall
        self.Rall=Rall
        self.Reall=Reall
        self.Call=Call  
         
        self.ser = servername
        self.db=db
        self.uid=UID
        self.pwd=PWD
        self.show()
        self.checkDB()
        self.bindWidget()
        self.dropdrow_setting_Name()
        self.insert_moudle()
       

#   driver=ODBC Driver 17 for SQL Server;server={self.ser};database={self.db};trusted_connection=yes;


    def checkDB(self):
     
        try:
            self.cursor.execute(
            """ DROP TABLE IF EXISTS [BigTable]
            SELECT S.StationName,Run.RunName,FRD.ObsTimestamp,P.PartName,F.FeatureName,FRD.Value
            INTO [BigTable]
            FROM Station AS S INNER JOIN Run AS Run
            ON S.StationID = Run.StationID
            INNER JOIN FeatureRunData AS FRD
            ON Run.RunID = FRD.RunID
            INNER JOIN Feature AS F
            ON FRD.FeatureID = F.FeatureID
            INNER JOIN Part AS P
            ON F.PartID = P.PartID;""")
            self.cursor.commit()
        except:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("เกิดข้อผิดพลาดในการดึงข้อมูล Database")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            self.close()

    def bindWidget(self):
          # self.dropdrow_setting()
            self.pushButton.clicked.connect(self.choose)
            self.ButtonTool.clicked.connect(self.time_c)
            self.ButtonTool_2.clicked.connect(self.save_setting)
            self.SheetBox_15.currentTextChanged.connect(self.partname)
            self.SheetBox_16.currentTextChanged.connect(self.feature_name)
            self.SheetBox_2.currentTextChanged.connect(self.insert_setting)
            self.ButtonTool_5.clicked.connect(self.search)
         
    def insert_moudle(self):
        self.cursor.execute(
            """SELECT DISTINCT StationName FROM [BigTable] """)
        name = self.cursor.fetchall()
        name = [name[0] for name in name[0:len(name)]]
        self.SheetBox_15.addItems(name)
        
        

        self.cursor.execute(
            """SELECT ObsTimestamp FROM [BigTable] """)
        datest = self.cursor.fetchall()

        datest = [datest[0] for datest in datest[0:len(datest)]]
        self.timestamp = datest
        dates = []
        for date in datest:
          
            dates.append(date.date())

        self.date = dates

    def featclear(self,partname):
        self.cursor.execute(
                    """SELECT DISTINCT FeatureName FROM [BigTable] WHERE PartName=? """, (partname))
        item = self.cursor.fetchall()
        item= [item[0] for item in item[0:len(item)]]
        item.append(' ')
        item.sort()
        for i in self.featurall:
            i.clear()
            i.addItems(item)
       
        return
    
    def insert_setting(self, name):
       
        self.SheetBox_16.blockSignals(True)
        for i in self.featurall:
            i.clear()
        self.interval.clear()
        colin = []
        rowend = []
        rowin = []
        c = ""
        if name != " ":
            self.cursor.execute(
                """SELECT * FROM USER_Setting_verti WHERE Name = ? """, name)
            data = self.cursor.fetchall()
            if data != []:
                rows = [data[1] for data in data[0:len(data)]]
                rows_end = [data[2] for data in data[0:len(data)]]
                cols_start = [data[3] for data in data[0:len(data)]]
                Fname = [data[4] for data in data[0:len(data)]]
                
                StationName = [data[5] for data in data[0:len(data)]]
                
                Partname = [data[6] for data in data[0:len(data)]]
                Date = [data[7] for data in data[0:len(data)]]
                Time_start = [data[8] for data in data[0:len(data)]]
                Time_end = [data[9] for data in data[0:len(data)]]
                interval=[data[10] for data in data[0:len(data)]]
                Fname=Fname[0].split("_._")
                col_start = cols_start[0].split(" ")
                row_end = rows_end[0].split(" ")
                row = rows[0].split(" ")
                Fnameall=[]
                try:
                    for x in col_start:
                           
                            x = x.split("_")
                           
                            colin.append(get_column_letter(int(x[0])))
                    for z in row_end:
                            z = z.split("_")
                            rowend.append(z[0])
                    for y in row:
                            y = y.split("_")
                            rowin.append(y[0])
                except:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setText(
                        "Error กรุณาตรวจสอบคอลัมน์และแถวใหม่ แล้วทำการ save setting ทับชื่อเดิม")
                    msg.setWindowTitle("Error")
                    msg.setStandardButtons(QMessageBox.Ok)
                    retval = msg.exec_()
                self.cursor.execute(
                    """SELECT DISTINCT FeatureName FROM [BigTable] WHERE PartName=? """, (Partname[0]))
                self.featclear(Partname[0])
                for i in Fname:
                    
                    Fnameall.append(i)
            
                self.interval.setText(interval[0])
                for i,v in enumerate(self.featurall) :              
                    
                    try:
                        v.setCurrentIndex(v.findText(Fnameall[i]))
                        self.Call[i].setText(colin[i])
                        self.Reall[i].setText(rowend[i])
                        self.Rall[i].setText(rowin[i])
                    except:
                        self.Call[i].setText(c)
                        self.Reall[i].setText(c)
                        self.Rall[i].setText(c)
                self.SheetBox_15.setCurrentIndex(
                        self.SheetBox_15.findText(StationName[0]))
                    
                self.SheetBox_16.setCurrentIndex(
                        self.SheetBox_16.findText(Partname[0]))
                self.calendarWidget.setSelectedDate(Date[0])
                self.timeEdit.setTime(Time_start[0])
                self.timeEdit_2.setTime(Time_end[0])        
     
        else:
                  
                    for i,v in enumerate(self.featurall) :              
                        v.setCurrentIndex(v.findText(" "))
                        self.Call[i].clear()
                        self.Reall[i].clear()
                        self.Rall[i].clear()
                  
        self.SheetBox_16.blockSignals(False)

    def partname(self,text):
        self.SheetBox_16.clear()
        self.cursor.execute(
            """SELECT DISTINCT PartName FROM [BigTable] WHERE StationName=? """,text)
        name = self.cursor.fetchall()
        name = [name[0] for name in name[0:len(name)]]
        name.append(' ')
        name.sort()
        self.SheetBox_16.addItems(name)
    
    
    def feature_name(self, name):
   
        text = self.SheetBox_2.currentText()
        # conn = pyodbc.connect(

        #     'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
        #     self.ser+'; DATABASE=MeasurLink9;trusted_connection=yes;'
        # )
        if name != " ":
        #         cursor = conn.cursor()
                self.featclear(name)
                for i,v in enumerate(self.featurall) :              
                        self.Call[i].clear()
                        self.Reall[i].clear()
                        self.Rall[i].clear()
                # self.SheetBox_2.currentTextChanged.connect(self.insert_setting)
                return

    def search(self):
        data=[]
        while True:
            self.cursor.execute(
                        """SELECT distinct [StationName]
                 ,[PartName]
                ,CAST(ObsTimestamp AS DATE)
               
               
                
                
            FROM [BigTable] """)

            datest = self.cursor.fetchone()
            if datest != None:
               
                data.append(datest)
                self.cursor.execute(
                            """;WITH CTE AS(SELECT  * FROM BigTable WHERE StationName=? AND [PartName]=? AND CAST(ObsTimestamp AS DATE)=?  ) DELETE FROM CTE """, (datest[0],datest[1],datest[2]))
            else:
                break
        self.checkDB()   
        self.window = QtWidgets.QMainWindow()
        self.table=Table(data)
      

    def choose(self):
            self.pushButton.setEnabled(False)
            Tk().withdraw()
            filename = askopenfilename()
            try:
                wb = load_workbook(filename)
                self.flieName.clear()
                self.flieName.setText(filename)
                self.SheetBox.clear()
                self.SheetBox.addItems(wb.sheetnames)
                
                self.location = filename
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("เปิดด้วยไฟล์ Excel เท่านั้น")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
            self.pushButton.setEnabled(True)

    def time_c(self):
        start = self.timeEdit.time()
        end = self.timeEdit_2.time()
        h_start = start.hour()
        m_start = start.minute()
        h_end = end.hour()
        m_end = end.minute()
        time_start = datetime.time(h_start, m_start, 0)
        time_end = datetime.time(h_end, m_end, 0)
        self.inputdb(time_start, time_end)

    def inputdb(self, time_start, time_end):
        value = self.calendarWidget.selectedDate()
        value = value.toPyDate()
        datetime_start = datetime.datetime.combine(value, time_start)
        datetime_end = datetime.datetime.combine(value, time_end)
        if value in self.date:
            try:
                textcom = str(self.SheetBox.currentText())
                filename = self.location
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("กรุณาเลือกไฟล์")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
                return
            start_row=[]
            featdata=[]  
            start_colunm=[]
            end_row=[]

            for i,v in enumerate(self.Rall):
                featdata.append(self.featurall[i].currentText())
                try:
                    startrow=int(v.text())
                    start_row.append(startrow)
                    end_row.append(int(self.Reall[i].text()))
                except:
                    start_row.append(0)
                    end_row.append(0)  
                try:
                    start_colunm.append(column_index_from_string(
                        str(self.Call[i].text()).upper()))
                  
                   
                except:
             
                    start_colunm.append(0)
                    
            StationName = self.SheetBox_15.currentText()
            
            PartName = self.SheetBox_16.currentText()
            try:
                interval=int(self.interval.text())
                if interval < 1 :
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Information)
                        msg.setText("ใส่ค่าInvervalตั้งแต่1ขึ้นไป")
                        msg.setStandardButtons(QMessageBox.Ok)
                        msg.setWindowTitle("Error")
                        retval = msg.exec_()
                        return
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("ใส่ค่าInvervalเป็นเลขจำนวนเต็ม")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
                return
            
           
            server = self.ser
            db=self.db
            uid=self.uid
            pwd=self.pwd
            self.window = QtWidgets.QMainWindow()
            self.textcom = str(self.SheetBox.currentText())
            self.ui2 = Choosepage_verti(self.window, filename, start_colunm, end_row, start_row, 
                                         textcom, featdata,  StationName,  PartName, value, server, datetime_start, datetime_end,interval,db,uid,pwd)
            self.window.show()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("ไม่มีข้อมูลในวันที่เลือก")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
    def dropdrow_setting_Name(self):
        self.SheetBox_2.clear()
        try:
            self.cursor.execute(
                """SELECT Name FROM USER_Setting_verti """)
            name = self.cursor.fetchall()

            name = [name[0] for name in name[0:len(name)]]
            name.append(" ")
            name.sort()
            self.SheetBox_2.addItems(name)
            self.SheetBox_2.setCurrentIndex(self.SheetBox_2.findText(' '))
        except:
            pass

    def save_setting(self):
        self.ButtonTool_2.setEnabled(False)
        i = 0
        saverow = []
        while 1:
                try:
                    self.cursor.execute("""
                        CREATE TABLE USER_Setting_verti (
                            Name       NVARCHAR (255)  PRIMARY KEY ,
                            Rstart         NVARCHAR (MAX),
                            Rend         NVARCHAR (MAX),
                            Cstart        NVARCHAR (MAX),
                            Fname     NTEXT,
                          
                            StationName      NVARCHAR (MAX),
                           
                            PartName      NVARCHAR (MAX),
                            Date        DATETIME,
                            Time_start TIME,
                            Time_end TIME,
                            Interval NVARCHAR (MAX),

                                            ) """)
                    self.conn.commit()

                except:
                    print('ตารางเดิมมีอยู่แล้ว')
                    break
        name = easygui.enterbox("ใส่ชื่อ")

        if name == None:
            self.ButtonTool_2.setEnabled(True)
            return name
        colcheck_start = []
        rowcheck_end = []
        
        for i,v in enumerate(self.Call):
                t=v.text().upper() 
                try:
                    t=str(column_index_from_string(t))+'_'+str(i)
                    if not t in colcheck_start:
                        colcheck_start.append(t)
                    else:
                        continue
                except:
                    break
        for i,v in enumerate(self.Reall):
            t=v.text()+'_'+str(i)
            rowcheck_end.append(t)
        for i,v in enumerate(self.Rall):
            t=v.text()+'_'+str(i)
            saverow.append(t)
        for i,v in enumerate(self.Rall):
            v.clear()
            self.Call[i].clear()
            self.Reall[i].clear()


        check = []
        featdata=[]
        for i in self.featurall:
            t=i.currentText()
        
            if t ==' ':
                continue
            featdata.append(i.currentText())
        StationName = self.SheetBox_15.currentText()
        
        PartName = self.SheetBox_16.currentText()
        value = self.calendarWidget.selectedDate()
        value = value.toPyDate()
        time_start = self.timeEdit.time()
        time_end = self.timeEdit_2.time()
        time_start = time_start.toPyTime()
        time_end = time_end.toPyTime()
        interval=self.interval.text()
     
        for y in saverow:
            if y.startswith('_'): 
                check.append(y)
        # หาค่าที่มีในcolcheck_startแต่ไม่มีในcheck https://www.youtube.com/watch?v=KtvYqmQNRfs
        saveroww = list(filter(lambda y: y not in check, saverow))
        i=0
        while i<20 :
            i=i+1
            try:
                try:
                
                    self.cursor.execute("""
                                                INSERT INTO USER_Setting_verti (Name,Cstart,Rend,Rstart,Fname,StationName,PartName,Date,Time_start,Time_end,Interval)
                                                VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (name, " ".join(str(x) for x in colcheck_start), " ".join(str(x) for x in rowcheck_end), " ".join(str(y) for y in saveroww),"_._".join(str(x) for x in featdata), StationName,  PartName, value,time_start,time_end,interval))
                    break
                except:
                
                    self.cursor.execute("""DELETE FROM USER_Setting_verti WHERE Name = ?;
                                        """, (name))
                    self.cursor.execute("""                          
                                            INSERT INTO USER_Setting_verti (Name,Cstart,Rend,Rstart,Fname,StationName,PartName,Date,Time_start,Time_end,Interval)
                                            VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (name, " ".join(str(x) for x in colcheck_start), " ".join(str(x) for x in rowcheck_end), " ".join(str(y) for y in saveroww),"_._".join(str(x) for x in featdata), StationName,  PartName, value,time_start,time_end,interval))
                    break
            except:
                    print('ตารางเดิมผิดพลาด ลบและสร้างตารางใหม่')
                    self.cursor.execute("""DROP TABLE IF EXISTS [USER_Setting_verti]""")
                    self.cursor.execute("""
                                CREATE TABLE USER_Setting_verti (
                                    Name       NVARCHAR (255)  PRIMARY KEY ,
                            Rstart         NVARCHAR (MAX),
                            Rend         NVARCHAR (MAX),
                            Cstart        NVARCHAR (MAX),
                            Fname     NTEXT,
                          
                            StationName      NVARCHAR (MAX),
                           
                            PartName      NVARCHAR (MAX),
                            Date        DATETIME,
                            Time_start TIME,
                            Time_end TIME,
                            Interval NVARCHAR (MAX),

                                                    ) """)
        self.conn.commit()

        self.SheetBox_2.clear()
        try:
            self.cursor.execute(
                """SELECT Name FROM USER_Setting_verti """)
            name = self.cursor.fetchall()

            name = [name[0] for name in name[0:len(name)]]
            name.append(" ")
            name.sort()
            self.SheetBox_2.addItems(name)
        except:
            pass
        self.ButtonTool_2.setEnabled(True)




class Horizontal(QtWidgets.QMainWindow, main.Ui_MainWindow):
    def __init__(self,servername,UID,PWD,db):
        super(Horizontal,self).__init__()
        uic.loadUi("UI/Filter_Data_H.ui", self)
        # กำหนดค่าให้mainwindowsให้ใช้ฟังใช่พวกpyqtได้
        # MainWindow = QtWidgets.QMainWindow()
        # self.ui = main.Ui_MainWindow()  # ดึงคลาสจากUI3เข้ามาใช้ในคลาสนี้
        # self.ui.setupUi(MainWindow)  # ผ่านค่าMainwindowsเข้าไปในคลาสsetupUi
        # # เรียกใช้ฟังชั่นต่าง ในคลาสsetupUiเข้ามาใช้ในคลาสนี้ได้
        # self.setupUi(self)
        try :
            self.conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                                  servername+'; DATABASE='+db+';trusted_connection=yes;' )
            self.cursor = self.conn.cursor()
            
            
        except:
            self.conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; SERVER='+servername+'; DATABASE='+db+';UID='+UID+';PWD='+PWD)
                
            self.cursor = self.conn.cursor()
        Call=[]
        Call.append(self.Cstart)
        Call.append(self.Cstart2)
        Call.append(self.Cstart3)
        Call.append(self.Cstart4)
        Call.append(self.Cstart5)
        Call.append(self.Cstart6)
        Call.append(self.Cstart7)
        Call.append(self.Cstart8)
        Call.append(self.Cstart9)
        Call.append(self.Cstart10)
        Call.append(self.Cstart11)
        Call.append(self.Cstart12)
        Call.append(self.Cstart13)
        Call.append(self.Cstart14)
        Call.append(self.Cstart15)
        Call.append(self.Cstart16)
        Call.append(self.Cstart17)
        Call.append(self.Cstart18)
        Call.append(self.Cstart19)
        Call.append(self.Cstart20)
        Ceall=[]
        Ceall.append(self.Cend)
        Ceall.append(self.Cend2)
        Ceall.append(self.Cend3)
        Ceall.append(self.Cend4)
        Ceall.append(self.Cend5)
        Ceall.append(self.Cend6)
        Ceall.append(self.Cend7)
        Ceall.append(self.Cend8)
        Ceall.append(self.Cend9)
        Ceall.append(self.Cend10)
        Ceall.append(self.Cend11)
        Ceall.append(self.Cend12)
        Ceall.append(self.Cend13)
        Ceall.append(self.Cend14)
        Ceall.append(self.Cend15)
        Ceall.append(self.Cend16)
        Ceall.append(self.Cend17)
        Ceall.append(self.Cend18)
        Ceall.append(self.Cend19)
        Ceall.append(self.Cend20)
        
        Rall=[]
        Rall.append(self.Rstart)
        Rall.append(self.Rstart2)
        Rall.append(self.Rstart3)
        Rall.append(self.Rstart4)
        Rall.append(self.Rstart5)
        Rall.append(self.Rstart6)
        Rall.append(self.Rstart7)
        Rall.append(self.Rstart8)
        Rall.append(self.Rstart9)
        Rall.append(self.Rstart10)
        Rall.append(self.Rstart11)
        Rall.append(self.Rstart12)
        Rall.append(self.Rstart13)
        Rall.append(self.Rstart14)
        Rall.append(self.Rstart15)
        Rall.append(self.Rstart16)
        Rall.append(self.Rstart17)
        Rall.append(self.Rstart18)
        Rall.append(self.Rstart19)
        Rall.append(self.Rstart20)
        featurall=[]
        featurall.append(self.SheetBox_001)
        featurall.append(self.SheetBox_002)
        featurall.append(self.SheetBox_003)
        featurall.append(self.SheetBox_004)
        featurall.append(self.SheetBox_005)
        featurall.append(self.SheetBox_006)
        featurall.append(self.SheetBox_007)
        featurall.append(self.SheetBox_008)
        featurall.append(self.SheetBox_009)
        featurall.append(self.SheetBox_010)
        featurall.append(self.SheetBox_011)
        featurall.append(self.SheetBox_012)
        featurall.append(self.SheetBox_013)
        featurall.append(self.SheetBox_014)
        featurall.append(self.SheetBox_015)
        featurall.append(self.SheetBox_016)
        featurall.append(self.SheetBox_017)
        featurall.append(self.SheetBox_018)
        featurall.append(self.SheetBox_019)
        featurall.append(self.SheetBox_020)
        self.featurall=featurall
        self.Rall=Rall
        self.Ceall=Ceall
        self.Call=Call   
        self.ser = servername
        self.db=db
        self.uid=UID
        self.pwd=PWD
        self.show()
        self.checkDB()
        self.bindWidget()
        self.dropdrow_setting_Name()
        self.insert_moudle()


#   driver=ODBC Driver 17 for SQL Server;server={self.ser};database={self.db};trusted_connection=yes;

    def checkDB(self):
        try:
            self.cursor.execute(
                """ DROP TABLE IF EXISTS [BigTable]
                SELECT S.StationName,Run.RunName,FRD.ObsTimestamp,P.PartName,F.FeatureName,FRD.Value
                INTO [BigTable]
                FROM Station AS S INNER JOIN Run AS Run
                ON S.StationID = Run.StationID
                INNER JOIN FeatureRunData AS FRD
                ON Run.RunID = FRD.RunID
                INNER JOIN Feature AS F
                ON FRD.FeatureID = F.FeatureID
                INNER JOIN Part AS P
                ON F.PartID = P.PartID;""")
            self.cursor.commit()
        except:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("เกิดข้อผิดพลาดในการดึงข้อมูล Database")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            self.close()

    def bindWidget(self):
           # self.dropdrow_setting()
            self.pushButton.clicked.connect(self.choose)
            self.ButtonTool.clicked.connect(self.time_c)
            self.ButtonTool_2.clicked.connect(self.save_setting)
            self.SheetBox_16.currentTextChanged.connect(self.feature_name)
            self.SheetBox_2.currentTextChanged.connect(self.insert_setting)
            self.ButtonTool_5.clicked.connect(self.search)
            self.SheetBox_15.currentTextChanged.connect(self.partname)
    def insert_moudle(self):
        self.cursor.execute(
            """SELECT DISTINCT StationName FROM [BigTable] """)
        name = self.cursor.fetchall()
        name = [name[0] for name in name[0:len(name)]]
        self.SheetBox_15.addItems(name)
        
    
      

        self.cursor.execute(
            """SELECT ObsTimestamp FROM [BigTable] """)
        datest = self.cursor.fetchall()

        datest = [datest[0] for datest in datest[0:len(datest)]]
        self.timestamp = datest
        dates = []
        for date in datest:
       
            dates.append(date.date())

        self.date = dates
    def featclear(self,partname):
        self.cursor.execute(
                    """SELECT DISTINCT FeatureName FROM [BigTable] WHERE PartName=? """, (partname))
        item = self.cursor.fetchall()
        item= [item[0] for item in item[0:len(item)]]
        item.append(' ')
        item.sort()
        for i in self.featurall:
            i.clear()
            i.addItems(item)
       
        return
    def insert_setting(self, name):
        self.SheetBox_16.blockSignals(True)
        for i in self.featurall:
            i.clear()
        self.interval.clear()
        colin = []
        colend = []
        rowin = []
        c = ""
        if name != " ":
            self.cursor.execute(
                """SELECT * FROM USER_Setting WHERE Name = ? """, name)
            data = self.cursor.fetchall()

            if data != []:

                cols_start = [data[1] for data in data[0:len(data)]]
                cols_end = [data[2] for data in data[0:len(data)]]
                rows = [data[3] for data in data[0:len(data)]]
                Fname = [data[4] for data in data[0:len(data)]]
                
                StationName = [data[5] for data in data[0:len(data)]]
                
                Partname = [data[6] for data in data[0:len(data)]]
                Date = [data[7] for data in data[0:len(data)]]
                Time_start = [data[8] for data in data[0:len(data)]]
                Time_end = [data[9] for data in data[0:len(data)]]
                interval=[data[10] for data in data[0:len(data)]]
                Fname=Fname[0].split("_._")
                col_start = cols_start[0].split(" ")
                col_end = cols_end[0].split(" ")
                row = rows[0].split(" ")
                Fnameall=[]
                try:
                    for x in col_start:
                        x = x.split("_")
                        colin.append(get_column_letter(int(x[0])))
                    for z in col_end:
                        z = z.split("_")
                        colend.append(get_column_letter(int(z[0])))
                    for y in row:
                        y = y.split("_")
                        rowin.append(y[0])
                    
                        
                except:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setText(
                        "Error กรุณาตรวจสอบคอลัมน์และแถวใหม่ แล้วทำการ save setting ทับชื่อเดิม")
                    msg.setWindowTitle("Error")
                    msg.setStandardButtons(QMessageBox.Ok)
                    retval = msg.exec_()
                self.featclear(Partname[0])
         
                for i in Fname:
                    
                    Fnameall.append(i)
           
                self.interval.setText(interval[0])
                for i,v in enumerate(self.featurall) :              
                    
                    try:
                        v.setCurrentIndex(v.findText(Fnameall[i]))
                        self.Call[i].setText(colin[i])
                        self.Ceall[i].setText(colend[i])
                        self.Rall[i].setText(rowin[i])
                    except:
                        self.Call[i].setText(c)
                        self.Ceall[i].setText(c)
                        self.Rall[i].setText(c)
                self.SheetBox_15.setCurrentIndex(
                        self.SheetBox_15.findText(StationName[0]))
                    
                self.SheetBox_16.setCurrentIndex(
                        self.SheetBox_16.findText(Partname[0]))
                self.calendarWidget.setSelectedDate(Date[0])
                self.timeEdit.setTime(Time_start[0])
                self.timeEdit_2.setTime(Time_end[0])        
     
        else:
                 
                    for i,v in enumerate(self.featurall) :              
                        v.setCurrentIndex(v.findText(" "))
                        self.Call[i].clear()
                        self.Ceall[i].clear()
                        self.Rall[i].clear()
                  
                  
        self.SheetBox_16.blockSignals(False)

    
    def partname(self,text):
        self.SheetBox_16.clear()
        self.cursor.execute(
            """SELECT DISTINCT PartName FROM [BigTable] WHERE StationName=? """,text)
        name = self.cursor.fetchall()
        name = [name[0] for name in name[0:len(name)]]
        name.append(' ')
        name.sort()
        self.SheetBox_16.addItems(name)
        
    def feature_name(self, name):
  
        text = self.SheetBox_2.currentText()
        if name != " ":
            

                self.featclear(name)
                for i,v in enumerate(self.featurall) :              
                        self.Call[i].clear()
                        self.Ceall[i].clear()
                        self.Rall[i].clear()
                # self.SheetBox_2.currentTextChanged.connect(self.insert_setting)
                return

    def search(self):
        data=[]
        while True:
            self.cursor.execute(
                        """SELECT distinct [StationName]
                 ,[PartName]
                ,CAST(ObsTimestamp AS DATE)
               
                
                
                
            FROM [BigTable] """)

            datest = self.cursor.fetchone()
            if datest != None:
               
                data.append(datest)
                self.cursor.execute(
                            """;WITH CTE AS(SELECT  * FROM BigTable WHERE StationName=? AND [PartName]=? AND CAST(ObsTimestamp AS DATE)=?  ) DELETE FROM CTE """, (datest[0],datest[1],datest[2]))
            else:
                break
        self.checkDB()      
        self.window = QtWidgets.QMainWindow()
        self.table=Table(data)
       

    def choose(self):
            self.pushButton.setEnabled(False)
            Tk().withdraw()
            filename = askopenfilename()
            try:
                wb = load_workbook(filename)
                self.flieName.clear()
                self.flieName.setText(filename)
                self.SheetBox.clear()
                self.SheetBox.addItems(wb.sheetnames)
                # self.textcom = str(self.SheetBox.currentText())
                self.location = filename

            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("เปิดด้วยไฟล์ Excel เท่านั้น")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
            self.pushButton.setEnabled(True)

    def time_c(self):
        start = self.timeEdit.time()
        end = self.timeEdit_2.time()
        h_start = start.hour()
        m_start = start.minute()
        h_end = end.hour()
        m_end = end.minute()
        time_start = datetime.time(h_start, m_start, 0)
        time_end = datetime.time(h_end, m_end, 0)
        self.inputdb(time_start, time_end)

    def inputdb(self, time_start, time_end):
        value = self.calendarWidget.selectedDate()
        value = value.toPyDate()
        datetime_start = datetime.datetime.combine(value, time_start)
        datetime_end = datetime.datetime.combine(value, time_end)
        if value in self.date:
            try:
                textcom = str(self.SheetBox.currentText())
                filename = self.location
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("กรุณาเลือกไฟล์")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
                return
            start_row=[]
            featdata=[]  
            start_colunm=[]
            end_colunm=[] 
            for i,v in enumerate(self.Rall):
                featdata.append(self.featurall[i].currentText())
                try:
                    startrow=int(v.text())
                    start_row.append(startrow)
                except:
                    start_row.append(0) 
                try:
                    start_colunm.append(column_index_from_string(
                        str(self.Call[i].text()).upper()))
                  
                    end_colunm.append(column_index_from_string(str(self.Ceall[i].text()).upper()))
                except:
               
                    start_colunm.append(0)
                    end_colunm.append(0)    
            
            StationName = self.SheetBox_15.currentText()
            
            PartName = self.SheetBox_16.currentText()
            try:
                interval=int(self.interval.text())
                if interval < 1 :
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Information)
                        msg.setText("ใส่ค่าInvervalตั้งแต่1ขึ้นไป")
                        msg.setStandardButtons(QMessageBox.Ok)
                        msg.setWindowTitle("Error")
                        retval = msg.exec_()
                        return
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText("ใส่ค่าInvervalเป็นเลขจำนวนเต็ม")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setWindowTitle("Error")
                retval = msg.exec_()
                return
            # try:
            #     start_colunm = self.Cstart.text()
            #     start_colunm = column_index_from_string(
            #         str(start_colunm).upper())
            #     end_colunm = self.Cend.text()
            #     end_colunm = column_index_from_string(str(end_colunm).upper())
            # except:
            #     start_colunm = 0
            #     end_colunm = 0
            # try:
            #     start_colunm2 = self.Cstart2.text()
            #     start_colunm2 = column_index_from_string(
            #         str(start_colunm2).upper())
            #     end_colunm2 = self.Cend2.text()
            #     end_colunm2 = column_index_from_string(
            #         str(end_colunm2).upper())
            # except:
            #     start_colunm2 = 0
            #     end_colunm2 = 0

            # try:
            #     start_colunm3 = self.Cstart3.text()
            #     start_colunm3 = column_index_from_string(
            #         str(start_colunm3).upper())
            #     end_colunm3 = self.Cend3.text()
            #     end_colunm3 = column_index_from_string(
            #         str(end_colunm3).upper())
            # except:
            #     start_colunm3 = 0
            #     end_colunm3 = 0
            # try:
            #     start_colunm4 = self.Cstart4.text()
            #     start_colunm4 = column_index_from_string(
            #         str(start_colunm4).upper())
            #     end_colunm4 = self.Cend4.text()
            #     end_colunm4 = column_index_from_string(
            #         str(end_colunm4).upper())
            # except:
            #     start_colunm4 = 0
            #     end_colunm4 = 0
            # try:
            #     start_colunm5 = self.Cstart5.text()
            #     start_colunm5 = column_index_from_string(
            #         str(start_colunm5).upper())
            #     end_colunm5 = self.Cend5.text()
            #     end_colunm5 = column_index_from_string(
            #         str(end_colunm5).upper())
            # except:
            #     start_colunm5 = 0
            #     end_colunm5 = 0

            server = self.ser
            db=self.db
            uid=self.uid
            pwd=self.pwd
            self.window = QtWidgets.QMainWindow()
            self.textcom = str(self.SheetBox.currentText())
         
            self.ui2 = Choosepage(self.window, filename, start_colunm, end_colunm, start_row
                                 , textcom, featdata, StationName,  PartName, value, server, datetime_start, datetime_end,interval,db,uid,pwd)
            # self.ui2.setupUi(self.window, filename, start_colunm, end_colunm, start_row, start_colunm2, end_colunm2,
            #                 start_row2, start_colunm3, end_colunm3, start_row3, start_colunm4, end_colunm4,
            #                 start_row4, start_colunm5, end_colunm5,
            #                 start_row5, textcom, Featname1, Featname2, Featname3, Featname4,
            #                 Featname5,StationName,RunName,PartName,value,server)
            self.window.show()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("ไม่มีข้อมูลในวันที่เลือก")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()

    def dropdrow_setting_Name(self):
        self.SheetBox_2.clear()
        # conn = pyodbc.connect(

        #     'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
        #     self.ser+'; DATABASE=MeasurLink9;trusted_connection=yes;')

        # cursor = conn.cursor()
        try:
            self.cursor.execute(
                """SELECT Name FROM USER_Setting """)
            name = self.cursor.fetchall()

            name = [name[0] for name in name[0:len(name)]]
            name.append(" ")
            name.sort()
            self.SheetBox_2.addItems(name)
            self.SheetBox_2.setCurrentIndex(self.SheetBox_2.findText(' '))
        except:
            pass

    def save_setting(self):
        self.ButtonTool_2.setEnabled(False)
        i = 0

        saverow = []
        while 1:
                try:
                    self.cursor.execute("""
                        CREATE TABLE USER_Setting (
                            Name       NVARCHAR (255)  PRIMARY KEY ,
                            Cstart         NVARCHAR (MAX),
                            Cend         NVARCHAR (MAX),
                            Rstart        NVARCHAR (MAX),
                            Fname      NTEXT,                          
                            StationName      NVARCHAR (MAX),
                            PartName      NVARCHAR (MAX),
                            Date        DATETIME,
                            Time_start TIME,
                            Time_end TIME,
                            Interval NVARCHAR (MAX)

                                            ) """)
                    self.conn.commit()

                except:
                    print('สร้างใหม่ไม่ได้เพราะมีตารางอยู่แล้ว')
                    break
        name = easygui.enterbox("ใส่ชื่อ")

        if name == None:
            self.ButtonTool_2.setEnabled(True)
            return name
        colcheck_start = []
        colcheck_end = []
     
        for i,v in enumerate(self.Call):
                t=v.text().upper() 
                try:
                    t=str(column_index_from_string(t))+'_'+str(i)
                    if not t in colcheck_start:
                        colcheck_start.append(t)
                    else:
                        continue
                except:
                    break
        for i,v in enumerate(self.Ceall):
            t=v.text().upper() 
            try:
                    t=str(column_index_from_string(t))+'_'+str(i)
                    if not t in colcheck_end:
                        colcheck_end.append(t)
                    else:
                        continue
            except:
                    break
        for i,v in enumerate(self.Rall):
            t=v.text()+'_'+str(i)
            saverow.append(t)
        for i,v in enumerate(self.Rall):
            v.clear()
            self.Call[i].clear()
            self.Ceall[i].clear()


        check = []
        featdata=[]
        for i in self.featurall:
            t=i.currentText()
         
            if t ==' ':
                continue
            featdata.append(i.currentText())
            
        StationName = self.SheetBox_15.currentText()
        
        PartName = self.SheetBox_16.currentText()
        value = self.calendarWidget.selectedDate()
        value = value.toPyDate()
        time_start = self.timeEdit.time()
        time_end = self.timeEdit_2.time()
        time_start = time_start.toPyTime()
        time_end = time_end.toPyTime()
        interval=self.interval.text()
        for y in saverow:
            if y.startswith('_'):
                check.append(y)
        saveroww = list(filter(lambda y: y not in check, saverow))
        i=0
        while i<20:
            i=i+1
            try:
                try:
                    
                    self.cursor.execute("""
                                            INSERT INTO USER_Setting (Name,Cstart,Cend,Rstart,Fname,StationName,PartName,Date,Time_start,Time_end,Interval)
                                            VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (name, " ".join(str(x) for x in colcheck_start), " ".join(str(x) for x in colcheck_end), " ".join(str(y) for y in saveroww),"_._".join(str(x) for x in featdata), StationName,  PartName, value,time_start,time_end,interval))
                    break
                except:
                
                    self.cursor.execute("""DELETE FROM USER_Setting WHERE Name = ?;
                                        """, (name))
                    self.cursor.execute("""
                                            INSERT INTO USER_Setting (Name,Cstart,Cend,Rstart,Fname,StationName,PartName,Date,Time_start,Time_end,Interval)
                                            VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (name, " ".join(str(x) for x in colcheck_start), " ".join(str(x) for x in colcheck_end), " ".join(str(y) for y in saveroww),"_._".join(str(x) for x in featdata), StationName,  PartName, value,time_start,time_end,interval))
                    break
            except:
                print('ตารางที่มีอยู่เดิมคลาดเคลื่อน ต้องลบตารางแล้วสร้างใหม่')
                self.cursor.execute("""DROP TABLE IF EXISTS [USER_Setting]""")
                self.cursor.execute("""
                            CREATE TABLE USER_Setting (
                                Name       NVARCHAR (255)  PRIMARY KEY ,
                                Cstart         NVARCHAR (MAX),
                                Cend         NVARCHAR (MAX),
                                Rstart        NVARCHAR (MAX),
                                Fname      NTEXT,                          
                                StationName      NVARCHAR (MAX),
                                PartName      NVARCHAR (MAX),
                                Date        DATETIME,
                                Time_start TIME,
                                Time_end TIME,
                                Interval NVARCHAR (MAX)

                                                ) """)
        self.conn.commit()

        self.SheetBox_2.clear()
        try:
            self.cursor.execute(
                """SELECT Name FROM USER_Setting """)
            name = self.cursor.fetchall()

            name = [name[0] for name in name[0:len(name)]]
            name.append(" ")
            name.sort()
            self.SheetBox_2.addItems(name)
        except:
            pass

        self.ButtonTool_2.setEnabled(True)    

class Table(QMainWindow):
    def __init__(self,data):
        super(Table, self).__init__()
        print("...")
        uic.loadUi("UI/table.ui", self)
        self.tableWidget.setRowCount(len(data))  # set number of rows
        # this is fixed for myTableWidget, ensure that both of your tables, sql and qtablewidged have the same number of columns
        self.tableWidget.setColumnCount(3)
        for k,i in enumerate(data):
            i[2]=i[2].strftime("%d %b %Y ")
          
            for a in range(0,3):
             
                self.tableWidget.setItem(k,a,QtWidgets.QTableWidgetItem(i[a]))
        
        self.show()
        
        
       
class LoadingScreen(QWidget):
    def __init__(self):
        super().__init__()
        self.setFixedSize(400,200)
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.CustomizeWindowHint)

        self.label_animation= QLabel(self)
        self.movie = QMovie('UI/Loading_2.gif')
        self.label_animation.setMovie(self.movie)
        
        
    def startloading(self):
        self.movie.start()
        self.show()

    def stoploading(self):
        self.movie.stop()
        self.close()


class Login_Ui(QMainWindow):
    def __init__(self):
        super(Login_Ui, self).__init__()
        print("...")
        uic.loadUi("UI/Export Data to Excel.ui", self)
        self.loading=LoadingScreen()
        
        
        self.show()
        self.pushButton.clicked.connect(self.test)
        self.insert_server()
    
    
    
    def insert_server(self):
        try:
            f = open("sname.txt")
            print(f)
            for i in f:
                print(i)
                if  i.startswith('servername: '):
                    y=re.findall('\S+',i)
                    print(y)
                    self.server.setText(y[1])
                elif i.startswith('database: '):
                    y=re.findall('\S+',i)
                    print(y)
                    self.pwd_2.setText(y[1])
        except:
            pass
    def test(self):
        global uid
        global pwd
        try:
            uid=self.uid.text()
            pwd=self.pwd.text()           
        except:
            uid=NONE
            pwd=NONE
        if self.radioButton_2.isChecked() :
            self.horizontal(uid,pwd)
        else:
            self.vertical(uid,pwd)
    def save_servername(self,server,database):
        server='servername: '+ server+ '\n'
        database='database: '+database+ '\n'
        L = [server,database]
        try:
                f = open("sname.txt", "w")
            
                
                f.writelines(L)  
                f.close()
        except:   
                f = open("sname.txt", "a")
                f.writelines(L)  
                f.close()
      
                

    def vertical(self,uid,pwd):
        self.loading.startloading()
        server=self.server.text()
        db=self.pwd_2.text()
        self.thread = BackgroundThread(server,uid,pwd,db)
        self.thread.start()
        self.thread.result.connect(self.go_verti)

    def go_verti(self,result):
        server_name = self.server.text()
        db=self.pwd_2.text()
        if result == 'good_windows':
            self.save_servername(server_name,db)
            self.window = QtWidgets.QMainWindow()
            self.loading.stoploading()
            self.next = Vertical(server_name,uid,pwd,db)
            
            # self.close()
        elif result == 'good_user':
            self.save_servername(server_name,db)
            self.window = QtWidgets.QMainWindow()
            self.loading.stoploading()
            self.next = Vertical(server_name,uid,pwd,db)
            
            # self.close()
        else:
            self.loading.stoploading()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("ไม่สามารถเชื่อมต่อ Database ได้")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()

    def horizontal(self,uid,pwd):
        self.loading.startloading()
        server=self.server.text()
        db=self.pwd_2.text()
        self.thread = BackgroundThread(server,uid,pwd,db)
        self.thread.start()
        self.thread.result.connect(self.go_hori)

    def go_hori(self,result): 
        server_name = self.server.text()
        db=self.pwd_2.text()
      
        if result == 'good_windows':
            self.save_servername(server_name,db)
            self.window = QtWidgets.QMainWindow()
            self.loading.stoploading()
            self.next = Horizontal(server_name,uid,pwd,db)
            
            # self.close()
        elif result == 'good_user':
            self.save_servername(server_name,db)
            self.window = QtWidgets.QMainWindow()
            self.loading.stoploading()
            self.next = Horizontal(server_name,uid,pwd,db)
            
            # self.close()
        else:
            self.loading.stoploading()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("ไม่สามารถเชื่อมต่อ Database ได้")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            


class BackgroundThread(QtCore.QThread, ):
    result = QtCore.pyqtSignal(str)
    def __init__(self, server_name,UID,PWD,db):
        super().__init__()
        self.name = server_name
        self.uid=UID
        self.pwd=PWD
        self.db=db
    def run(self): 
        try:
            conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                                  self.name+'; DATABASE='+self.db+';trusted_connection=yes;')
            self.result.emit('good_windows')
        except:
            try:
                
                conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; SERVER='+self.name+';PORT=1433; DATABASE='+self.db+';UID='+self.uid+';PWD='+self.pwd)
                print(self.uid,self.pwd)
               
                
                self.result.emit('good_user')
            except:
                self.result.emit('bad') 

class Thread_regist(QtCore.QThread, ):
    result = QtCore.pyqtSignal(str)
    def __init__(self, name,company,email):
        super().__init__()
   
   
        self.name=name
        self.company=company
        self.email=email
    def run(self):
        try:
            yag = yagmail.SMTP(user=#mailคนส่ง,password='akapongalone')
            reciver = #เมล์คนรับ
            subject = 'Serial Key ของบริษัท'+ self.company
            body= f'บริษัท {self.company} \n\nพนักงานชื่อ {self.name} \n\n Emailบริษัท {self.email}\n\n Serial Key คือ {key_id}'
            yag.useralias='บริษัท '+ self.company
            yag.send(to=reciver,subject=subject,contents=[body])
            self.result.emit('ok')
        except:
            self.result.emit('error')
        
        
        
class Register(QMainWindow):
    def __init__(self):
        super(Register, self).__init__()
        uic.loadUi("UI/Register.ui", self)
        self.loading=LoadingScreen()
        self.show()
        self.pushButton.clicked.connect(self.register)
    def register(self):
        self.pushButton.setEnabled(False)
        name = self.name.text()
        company=self.corp.text()
        email=self.email.text()
        self.loading.startloading()
        self.thread = Thread_regist(name,company,email)
        self.thread.start()
        self.thread.result.connect(self.finish)         
    def finish(self,result):
        if result == 'ok' : 
            self.loading.stoploading() 
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("สำเร็จ")
            msg.setWindowTitle("Sucessful")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            self.close()
        elif result =='error':
            self.loading.stoploading() 
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("Error")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            self.close()
class Wellcome(QMainWindow):
    def __init__(self):
        super(Wellcome, self).__init__()
        uic.loadUi("UI/Serial Key.ui", self)
        self.show()
        self.pushButton_2.clicked.connect(self.activate)
        self.pushButton.clicked.connect(self.go_register)
        
    def go_register(self):
   
        self.window = QtWidgets.QMainWindow()
        self.go = Register()
        # self.close()
      
        
    def activate(self):
        insert_key=self.key.text()
        if insert_key == key_id:
            try:
                f = open("lisene.txt", "w")
                l=[key_id,"\n",str(date.today())]
                f.writelines(l)
                f.close()
            except:
                f = open("lisene.txt", "a")
                f.writelines(l)
                f.close()
            self.window = QtWidgets.QMainWindow()
            self.next = Login_Ui()
            self.close()
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("Serial Key ไม่ถูกต้อง")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            
       

check=True
while check:
    try :
        f = open("lisene.txt", "r")
        info=f.readlines()
        key_check=info[0]
        day_register=info[1]
        day_register=date.fromisoformat(day_register)
        timedif=date.today()-day_register
        timedif=timedif.days
    
        if not timedif > 30 :
            
            key_check=key_check.rstrip("\n")
            print(key_check,'hello wtf')
            print(key_check == key_id)
            if key_check == key_id:
                
                app = QApplication(sys.argv)
                            
                maina = Login_Ui()
                app.exec_()
            
            else:
              
                os.remove("lisene.txt")
        else :
            app = QApplication(sys.argv)
            maina = Wellcome()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("หมดเวลาทดลองใช้งาน")
            msg.setWindowTitle("Warning")
            msg.setStandardButtons(QMessageBox.Ok)
            retval = msg.exec_()
            app.exec_()
            
            os.remove("lisene.txt")
            break
    except:
        app = QApplication(sys.argv)
        maina = Wellcome()
        app.exec_()
        break



# if __name__ == "__main__":
#     app = QtWidgets.QApplication(sys.argv)
#     maina = Horizontal()
#     sys.exit(app.exec_())
#pyuic5 Select_Data.ui > select_h.py

# "DESKTOP-H45D3IV" MeasurLink9

# console=False , icon='icon.ico',
