# from django.http.response import JsonResponse
from typing import Type
from django.shortcuts import render,redirect
from django.http import HttpResponse,JsonResponse
from django.contrib.auth.models import User
from django.contrib import messages
# from .models import Processname,Work
from django.forms.models import model_to_dict
from django.views.decorators.csrf import csrf_exempt
# from rest_framework.parsers import JSONParser
from django.conf import settings
from django.core.files.storage import FileSystemStorage

from main.models import Model as md
from main.models import Tool as tl
from main.models import Tooltype as ty
from main.models import Work as wk
from main.models import Pattern as pt
# from cal_data.serializers import Serializer
from contextlib import suppress
import xlwings as xl
import pyodbc
import datetime
import time
from datetime import date ,datetime,timedelta
import re 
import statistics
from ast import literal_eval
import json
import os
import pandas as pd
import psutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
from xlwings.utils import rgb_to_int
import pythoncom
import win32com
import win32com.client

#####เลือกDatabase#######################
conn = pyodbc.connect(
                    'DRIVER={ODBC Driver 17 for SQL Server};SERVER=AKAPONG\SQLEXPRESS; DATABASE=S-pack;trusted_connection=yes;'
                )
cursor = conn.cursor()

###################หน้าแรก###########################
def hello(request):
    if request.method=='POST':
        try:
            sel_model=request.POST['model']
            sel_partname=request.POST['partname']
            partname =md.objects.values('partname').distinct()
            pattern=md.objects.filter(partname=sel_partname,model=sel_model).values('pattern').distinct()
            model =md.objects.filter(partname=sel_partname).values('model').distinct()
            pattern_list=list(pattern)
            print('kuyy',pattern_list[0]['pattern'])
            shift=pt.objects.filter(name=pattern_list[0]['pattern']).values('name_shift')
            print(shift)
            return render(request,'home.html',{'partname':partname,'sel_partname':sel_partname,'model':model,'shift':shift,'sel_model':sel_model,'pattern':pattern_list[0]['pattern']})
        except:
            sel_partname=request.POST['partname']
            partname =md.objects.values('partname').distinct()
            model =md.objects.filter(partname=sel_partname).values('model').distinct()
    
        return render(request,'home.html',{'partname':partname,'sel_partname':sel_partname,'model':model})
        
    
    partname =md.objects.values('partname').distinct() 

    print(partname,partname.query)
    return render(request,'home.html',{'partname':partname})


def tool_list():
    toolall=tl.objects.values('toolname','ttid')
    tooltype=ty.objects.values('ttid','type')
    return toolall,tooltype

def add(request):
    eiei=tool_list()
    pattern=pt.objects.values('name').distinct()
    
    return render(request,'add/model.html',{'tool':eiei[0],'type':eiei[1],'pattern':pattern})


def get_model(request):
    if request.is_ajax():
        print('axious')
        excel_file = request.FILES["file"]
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)
        return JsonResponse( {'sheet':sheets})
    global conn
    global cursor
    global cstart
    global cend
    col=[]
    row=[]
    sheets=request.POST.getlist('sheets')
    try:
        # print(sheets[0])
        sheets=literal_eval(sheets[0])
    except:
        pass
    
    try:
        
        uploaded=request.FILES['choose_file'] 
        wb = openpyxl.load_workbook(uploaded)
        sheets=wb.sheetnames
        excel=uploaded.name

        
        if not os.path.exists("C:/excel_test/Template_excel/"+excel):
            fs=FileSystemStorage("C:/excel_test/Template_excel")
            fs.save(uploaded.name,uploaded)
    except:
        
        excel=request.POST['exceltest']
    try:
        img_uploaded=request.FILES['img_file']
        img=img_uploaded.name
        if not os.path.exists("C:/excel_test/Add_img/"+img):
            fs2=FileSystemStorage("C:/excel_test/Add_img")
            fs2.save(img_uploaded.name,img_uploaded)
    except:
        img=request.POST['img']
    model=request.POST['model']
    machname=request.POST['machname']
    machno=request.POST['machno']
    machpro=request.POST['machpro']
    process=request.POST['process']
    partname=request.POST['partname']
    tool=request.POST['tool']
    sel_pattern=request.POST['pattern']
    
    tooltype=int(tool[0])
    toolsel=tool[1:]
    if tooltype == 1:
        tooltype='gonogo'
    elif tooltype == 2:
        tooltype='value'
    elif tooltype==3:
        tooltype='OK/NG'
    else :print(tooltype)
   
    point=request.POST.getlist('point')
    if tooltype == 'value':
        min=request.POST.getlist('min')
        max=request.POST.getlist('max')
    
    # print(img)
    sheet=request.POST['sheettest'] 
    
    if  sheet == '':

        sheet=request.POST['sheet']
    
    direction=pt.objects.filter(name=sel_pattern).values('direction').distinct()
 
    # print('di',len(direction[0]['direction']),direction[0] == 'horizontal')
    if direction[0]['direction'] == 'vertical':
        try:
            rstart_old=md.objects.filter(model=model,excel=excel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,pattern=sel_pattern,partname=partname).values('rstart').order_by('-rstart')[0:1]
            rstart_new=rstart_old[0]
            rstart_new=rstart_new['rstart']  
            rstart_new=int(rstart_new)+1
        except:
            rstart_new= pt.objects.filter(name=sel_pattern).values('rstart')
            rstart_new=int(rstart_new[0]['rstart'])

        for i, p in enumerate(point):
            shift=pt.objects.filter(name=sel_pattern).values('name_shift')
            shift=list(shift)
            # print(shift)
            if tooltype == 'value':
                for k in shift:
                        cstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('cstart')
                  
                        cstart=int(cstart[0]['cstart'])
                        md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,cstart=cstart,rstart=rstart_new+i,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],min=min[i],max=max[i],partname=partname)
                        # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',min=min[i],max=max[i],partname=partname)
            else:
                for k in shift:
                    # print(sel_pattern,k['name_shift'])
                    cstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('cstart')
                    # print(cstart)
                    cstart=int(cstart[0]['cstart'])
                    md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],partname=partname,cstart=cstart,rstart=rstart_new+i)
                # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',partname=partname)
    elif direction[0]['direction'] == 'horizontal':
        try:
            cstart_old=md.objects.filter(model=model,excel=excel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,pattern=sel_pattern,partname=partname).values('cstart').order_by('-cstart')[0:1]
        
            cstart_new=cstart_old[0]
            cstart_new=cstart_new['cstart']  
             
            cstart_new=int(cstart_new)+1
        except:
            cstart_new= pt.objects.filter(name=sel_pattern).values('cstart')
            cstart_new=int(cstart_new[0]['cstart'])
            
        for i, p in enumerate(point):
            shift=pt.objects.filter(name=sel_pattern).values('name_shift')
            shift=list(shift)
            # print(shift)
            if tooltype == 'value':
                for k in shift:
                        rstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('rstart')
                  
                        rstart=int(rstart[0]['rstart'])
                        md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,rstart=rstart,cstart=cstart_new+i,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],min=min[i],max=max[i],partname=partname)
                        # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',min=min[i],max=max[i],partname=partname)
            else:
                for k in shift:
                    rstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('rstart')       
                    rstart=int(rstart[0]['rstart'])
                    md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],partname=partname,rstart=rstart,cstart=cstart_new+i)
    eiei=tool_list()
    toolall=eiei[0]
    type_list=eiei[1]
    pattern=pt.objects.values('name').distinct()
    
    # print(sel_pattern)
   
    return render(request,'add/model.html', {'model':model,'toolsel':toolsel,'excel':excel,'machname':machname,'machno':machno,'tool':toolall,'type':type_list,'img':img,'machpro':machpro,'process':process,'sheet':sheet,'sheets':sheets,'partname':partname,'pattern':pattern,'sel_pattern':sel_pattern}) 


#######################เมื่อกดStart########################################
def main(request):          
            global excel
            global locate
            shift=request.POST['shift']
            pattern=request.POST['pattern']
            # if shift == 'A':
            #     shift='Day'
            # elif shift == 'B':
            #     shift='Night'
            # else:
            #     print(shift)
            partname_sel=request.POST['partname_rev']
            # print('hellloooooo',partname_sel)
            model_sel=request.POST['model']

           
            
            today=gen_today(pattern,shift)
            man_regis=request.POST['man_regis']
            time_in = datetime.now().time()
            data_model=md.objects.filter(model=model_sel,shift=shift,partname=partname_sel).values('model','point','toolname','shift','machine_line','machine_no','machine_process','processname','partname','wid')
            
            partno=request.POST['partno']
            print(data_model,'datamodel')
            for i in data_model:
                wk.objects.create(**i,date=today,partno=partno)
                #ค่าiเป็นแบบdictจึงต้องกำหนดตัวแปรให้เป็นแบบkwargโดยการเติม**
           
            wrid=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,result_gonogo__isnull=True,result__isnull=True,shift=shift,partno=partno).values('wid','rid').order_by('wid')[0:1]
            try :
                wid = wrid[0]['wid']
                rid = wrid[0]['rid']
                # print(wid,rid,type(wrid)) 
            except: 
                return check(request)  
            test=md.objects.filter(wid=wid).values('wid','min','max','cstart','rstart','img','point','toolname','tooltype','processname','sheet','excel','partname','machine_no')       
            vallist=test[0]
            vallist['rid']=rid
            
            count_wait=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=True,timeout__isnull=True,shift=shift,partno=partno).values('rid') 
            count_finish=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=False,timeout__isnull=False,shift=shift,partno=partno).values('rid')
            count_all=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,shift=shift,partno=partno).values('rid')
            count_wait=list(count_wait)
            count_all=list(count_all)
            count_finish=list(count_finish)
            percent=len(count_finish)/len(count_all) *100   
            percent=float ("{:.2f}".format(percent))
            # print(vallist['tooltype'])
            return render(request,'third version/insert.html',{'wait':len(count_wait),'finish':len(count_finish),'percent':percent,'all':len(count_all),'model':model_sel,'vallist':vallist,'timein':str(time_in),'tooltype':vallist['tooltype'],'today':str(today),'man_regis':man_regis,'shift':shift,'partname':partname_sel,'pattern':pattern,'partno':partno})
    

#######################ฟังก์ชั่นไว้รับค่าแล้วประมวลผล####################################   
@csrf_exempt
def main_2(request): 
        global time_out
        partno=request.POST['partno']
        shift=request.POST['shift']
        partname_sel=request.POST['partname']
        rid=request.POST['rid']
        wid=request.POST.getlist('wid')
        sheet=request.POST['sheet']
        model_sel=request.POST['model']
        # vaid='value_'+str(wid)
        # location=request.POST['location']
        tool_select=request.POST['tool']
        pattern=request.POST['pattern']
        # location=literal_eval(location)     #แปลงString '[sdgd]' เป็นlist [sdgd]
        man=request.POST['man_regis']
        excel=request.POST['excel'] 
        cstart=request.POST['cstart']
        rstart=request.POST['rstart']
        tooltype_sel=request.POST['tooltype']
        result=request.POST['value']
        time_in=request.POST['timein']
        sel_point=request.POST['point']
        today=request.POST['today']
        today=date.fromisoformat(today)
        location=file_check(model_sel,partname_sel, excel,shift,pattern,partno)

        time_out = datetime.now().time()
        timein_time = datetime.strptime(time_in,'%H:%M:%S.%f').time() 
        duration = datetime.combine(today, time_out) - datetime.combine(today, timein_time)    #หาผลต่างเวลาเข้า-ออก
       
        duration=duration.total_seconds()
        
        print(tooltype_sel=='gonogo',(tooltype_sel))
        if str(tooltype_sel) == 'value':
            wk.objects.filter(rid=rid).update(point=sel_point,man=man,duration=duration,result=result,timein=time_in,timeout=time_out,shift=shift)
            
        elif str(tooltype_sel) == 'gonogo' :
            wk.objects.filter(rid=rid).update(point=sel_point,man=man,duration=duration,result_gonogo=result,timein=time_in,timeout=time_out,shift=shift)
            
            
        else:print(tooltype_sel=='gonogo',(tooltype_sel))
        
        
        
        
         
            
        
            
    #     cursor.execute("""SELECT  Top 1 [Timeout]
    #     ,[Duration]
    # FROM [Kawazaki].[dbo].[Work] WHERE rid=? AND Date=CAST(GETDATE() AS Date) Order by Timeout DESC AND wi
    #                                 ;""",(rid))
            # time=cursor.fetchall() 
            # time=[time[1] for time in time[0:len(time)]]
            # time=float(time[0])/60
        
            
            
        all_mean=[]
        avg=wk.objects.filter(rid=rid).values_list('duration',named=False).distinct()
        # cursor.execute("""SELECT DISTINCT [Duration]
        #                             FROM Work
        #                             WHERE rid=?;""",(rid))
        # avg=cursor.fetchall() 
        # avg=[avg[0] for avg in avg[0:len(avg)]]
        print(avg)
        # try:
        #                 # avg =list(map(int, avg))
        #                 print(avg)
        #                 time=sum(avg)/60
        #                 time=float ("{:.2f}".format(time))
        #                 # mean=statistics.mean(avg)  
        #                 # all_mean.append(mean)
        # except:print('pass')
        # try:
               
        #         all_mean=statistics.mean(all_mean)
        # except:all_mean=0
        
            
        t = True
        while t:
                wb = xl.Book(location)
                print(sheet)
                time.sleep(0.5)
                ws = wb.sheets[sheet]  
                try: 
                    ws.activate() 
                except: 
                    pass
                # wb.app.activate(steal_focus=True)
                # ws.cells(2,5).value = str(partname_sel)
                ws.cells(int(rstart),int(cstart)).api.Font.Bold=True
                duration=wk.objects.filter(rid=rid).values('duration')
                duration=duration[0]['duration']
                print('pattern',pattern)
                data=pt.objects.filter(name=pattern,name_shift=shift).values('shift_type','timeend')
                print(data,'daa')
                shift_c = data[0]['shift_type']
                if shift_c == 'Day':
                    ct=16
                elif shift_c=='Night':
                    ct=17    
                if tooltype_sel == 'value':
                    # minmax=wk.objects.filter(rid=rid).values('min','max')
                    # min=minmax[0]['min']
                    # max=minmax[0]['max']
                    # print(result)
                    # if float(result) < min or float(result) > max:
                    
                    #     ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    #     ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((136,8,8))
                    #     ws.cells(int(rstart),int(ct)).value = duration       
                    # else :
                        ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                        ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((0,0,0))
                        # ws.cells(int(rstart),int(ct)).value = duration 
                elif result == 'nogo' :
                    result='NG'
                    ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((136,8,8))
                    # ws.cells(int(rstart),int(ct)).value = duration 
               
                else:
                    result='OK'
                    ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((0,0,0))
                    # ws.cells(int(rstart),int(ct)).value = duration 
                    
                # ws.cells(int(rtime[0]),int(ctime[0])-1).value = all_mean
                # ws.cells(int(rstart[0]),int(cstart[0])+i).value = str(result[1])
                # ws.cells(int(rstart),int(cstart)+6).value = avg[0]
                wb.save()
                t=False
             
#DELETE FROM Work WHERE Timeout NOT IN ( SELECT MAX(Timeout) Timeout FROM Work Where wid = 7)
        
##################################################################################################################################################################
        vallist={} 
           
            
        time_in = datetime.now().time()
        today = gen_today(pattern,shift)
        print('kk',model_sel,partname_sel,shift)   
        wrid=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,result_gonogo__isnull=True,result__isnull=True,shift=shift,partno=partno).values('wid','rid').order_by('wid')[0:1]
        print(wrid)
        try :
            wid = wrid[0]['wid']
            rid = wrid[0]['rid']
                 
        except: 
            return check(request)  
        test=md.objects.filter(wid=wid).values('wid','min','max','cstart','rstart','img','point','toolname','tooltype','processname','sheet','excel','partname','machine_no')       
        vallist=test[0]
        vallist['rid']=rid
        print(vallist['sheet'],'next sheet')
        ws_next=wb.sheets[vallist['sheet']]
        ws_next.activate() 
        count_wait=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=True,timeout__isnull=True,shift=shift,partno=partno).values('rid') 
        count_finish=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=False,timeout__isnull=False,shift=shift,partno=partno).values('rid')
        count_all=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,shift=shift,partno=partno).values('rid')
        count_wait=list(count_wait)
        count_all=list(count_all)
        count_finish=list(count_finish)
        percent=len(count_finish)/len(count_all) *100   
        percent=float ("{:.2f}".format(percent))
        print(vallist)
       
        return render(request,'third version/insert.html',{'shift':shift,'wait':len(count_wait),'finish':len(count_finish),'percent':percent,'all':len(count_all),'partname':partname_sel,'tool_select':tool_select,'tooltype':vallist['tooltype'],'tool':tool_select,'vallist':vallist,'timein':str(time_in),'man_regis':man,'location_next':location,'model':model_sel,'today':str(today),'pattern':pattern,'partno':partno})
        # time_in= datetime.now().strftime('%H:%M:%S')

def skip(request):
        excel=request.POST['excel'] 
        partno=request.POST['partno']
        model_sel=request.POST['model']
        shift=request.POST['shift']
        rid=request.POST['rid']
        man=request.POST['man_regis']
        tooltype_sel=request.POST['tooltype']
        sel_point=request.POST['point']
        partname_sel=request.POST['partname']
        pattern=request.POST['pattern'] 
        sheet=request.POST['sheet']
        location=file_check(model_sel,partname_sel, excel,shift,pattern,partno)
        wb = xl.Book(location)
          
        ws = wb.sheets[sheet]   
        ws.activate() 
        if str(tooltype_sel) == 'value':
            wk.objects.filter(rid=rid).update(point=sel_point,man=man,result='skip',shift=shift)
            
        elif str(tooltype_sel) == 'gonogo' :
            wk.objects.filter(rid=rid).update(point=sel_point,man=man,result_gonogo='skip',shift=shift)
            
            
        else:print(tooltype_sel=='gonogo',(tooltype_sel))
        
        vallist={} 
           
            
        time_in = datetime.now().time()
        today = gen_today(pattern,shift)
        print('kk',model_sel,partname_sel,shift)   
        wrid=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,result_gonogo__isnull=True,result__isnull=True,shift=shift).values('wid','rid').order_by('wid')[0:1]
        print(wrid)
        try :
            wid = wrid[0]['wid']
            rid = wrid[0]['rid']
                 
        except: 
            return check(request)  
        test=md.objects.filter(wid=wid).values('wid','min','max','cstart','rstart','img','point','toolname','tooltype','processname','sheet','excel','partname','machine_no')       
        vallist=test[0]
        vallist['rid']=rid
        ws_next=wb.sheets[vallist['sheet']]
        ws_next.activate() 
        
        count_wait=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=True,timeout__isnull=True,shift=shift,partno=partno).values('rid') 
        count_finish=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,timein__isnull=False,timeout__isnull=False,shift=shift,partno=partno).values('rid')
        count_all=wk.objects.filter(model=model_sel,date=today,partname=partname_sel,shift=shift,partno=partno).values('rid')
        count_wait=list(count_wait)
        count_all=list(count_all)
        count_finish=list(count_finish)
        percent=len(count_finish)/len(count_all) *100   
        percent=float ("{:.2f}".format(percent))
        
        return render(request,'third version/insert.html',{'shift':shift,'wait':len(count_wait),'finish':len(count_finish),'percent':percent,'all':len(count_all),'partname':partname_sel,'tooltype':vallist['tooltype'],'vallist':vallist,'timein':str(time_in),'man_regis':man,'model':model_sel,'today':str(today),'pattern':pattern,'partno':partno})
########################################หน้า Overall ############################
def check(request):
        shift=request.POST['shift']
        pattern=request.POST['pattern']
        partno=request.POST['partno']
        try:
            partname_sel=request.POST['partname']
        except:
            partname_sel=request.POST['partname_select']
        model=request.POST['model']                    
        man=request.POST['man_regis']
        today = gen_today(pattern,shift)
        # if shift == 'A':
        #             shift='Day'
        # elif shift == 'B':
        #             shift='Night'
            
        print(today,shift,partname_sel,model,man)
        old_data=wk.objects.filter(timein__isnull=False,timeout__isnull=False,date=today,model=model,partname=partname_sel,shift=shift,partno=partno).values('rid','toolname','result_gonogo','machine_no','result','point','partname','wid','model')
        old_data=list(old_data)
       
        data=[]
        for i in old_data:
            wid=i['wid']
            sheet=md.objects.filter(wid=wid).values('sheet')  
            sheet=sheet[0]['sheet']
            i['sheet']=sheet
            data.append(i) 
           
        wid_c=wk.objects.filter(model=model,date=today,partname=partname_sel,result_gonogo__isnull=True,result__isnull=True,shift=shift,partno=partno).values('wid').order_by('wid')[0:1]
        try:
                x=wid_c[0]
                finish=None
        except:
                finish='Finish'
        return render (request,'third version/check.html',{'man_regis':man,'shift':shift,'alllist':data,'model':model,'partname':partname_sel,'finish':finish,'pattern':pattern,'partno':partno})  


def file_check(model,Partname,excel,shift,pattern,partno):
        today = gen_today(pattern,shift)
        directory = Partname+'_'+str(today)
        parent_dir = "C:/excel_test/save file/"
        path = os.path.join(parent_dir, directory)
        model= re.sub(r"[^a-zA-Z0-9]", "_", model)
        partno= re.sub(r"[^a-zA-Z0-9]", "_", partno)
        try:         
                os.mkdir(path)
        except:pass
        locate={}
        try:
                    # e=excel.split('.')
            
            location=(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx") 
            print(location)
            wb = xl.Book(location)
                        
        except FileNotFoundError:
                        e=excel.split('.')
                        location=(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx") 
                      
                        location = 'C:/excel_test/Template_excel/'+excel
                        closeexcel(location)
                        wb = xl.Book(location)
                        
                        print(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx")
                        wb.save(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx") 
                        location=(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx") 
        except pythoncom.com_error as error :  
                        # closeexcel(location)        
                        location=(path+'/'+Partname+'_'+partno+'_'+model+'_'+str(today)+".xlsx")  
                     
        print('eiei',location)                
        return location  

def gen_today(pattern,shift):
    print('gentoday',shift)
    data=pt.objects.filter(name=pattern,name_shift=shift).values('shift_type','timeend')
    shift = data[0]['shift_type']
    print(shift)
    timeend=data[0]['timeend']
    timeend=timeend.split(':')
    hour=int(timeend[0])
    minute=int(timeend[1])
    if shift == 'Night':
                now = datetime.now()
                today0am = now.replace(hour=0, minute=0, second=0, microsecond=0)
                today7am = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if now >= today0am and now<= today7am:
                    today = date.today() - timedelta(days=1)
                else:today = date.today()
    else:today = date.today()
    return today

def closeexcel(location):
    from win32com.client import Dispatch

    excel = Dispatch("Excel.Application")
    excel.Visible = False
    workbook = excel.Workbooks.Open(location)

    # with saving
    excel.DisplayAlerts = False

    excel.ActiveWorkbook.Save()
    excel.Quit()

    #without saving

    map(lambda book: book.Close(False), excel.Workbooks)
    excel.Quit()    
def edit(request): 
    partno=request.POST['partno']
    tooltype=request.POST['tooltype']
    rid=request.POST['rid']
    result=request.POST['value']
    Partname=request.POST['partname']
    wid=request.POST['wid']
    shift=request.POST['shift']
    model=request.POST['model']
    pattern=request.POST['pattern']
    print('result',result)
    data=md.objects.filter(wid=wid).values('excel','sheet','cstart','rstart','tooltype')
    # cursor.execute("""SELECT  Excel,Sheet,Cstart,Rstart
                                
    #                         FROM [Model]

    #                         Where  wid=?""",wid)
    # data = cursor.fetchall()
    excel=data[0]['excel']
    sheet=data[0]['sheet']
    cstart=data[0]['cstart']
    rstart=data[0]['rstart']
    tooltype=data[0]['tooltype']
    
    location=file_check(model,Partname, excel,shift,pattern,partno) 
    
    print('tooltype',tooltype)
    if tooltype == 'gonogo':
        # cursor.execute("""UPDATE Work SET Result_gonogo =?
        #                     Where  rid=?""",result,rid)
        
        # cursor.commit()
        wk.objects.filter(rid=rid).update(result_gonogo=result)
    elif tooltype == 'value':
        # cursor.execute("""UPDATE Work SET Result =?
        #                     Where  rid=?""",result,rid)
        # cursor.commit()
        wk.objects.filter(rid=rid).update(result=result)
    else:
        return 'x'
    t = True
    while t:
        
                wb = xl.Book(location)
                ws = wb.sheets[sheet]     
                ws.activate() 
                # ws.cells(2,5).value = str(Partname)
                # duration=wk.objects.filter(rid=rid).values('duration')
                # duration=duration[0]['duration']
                data=pt.objects.filter(name=pattern,name_shift=shift).values('shift_type','timeend')
                shift = data[0]['shift_type']
                
                if tooltype == 'value':
                    # minmax=wk.objects.filter(rid=rid).values('min','max')
                    # # cursor.execute("""SELECT Min,Max
                    # #                 FROM Work
                    # #                 WHERE rid=?;""",(rid))
                    # # data=cursor.fetchall() 
                    # min=minmax[0]['min']
                    # max=minmax[0]['max'] 
                    # if float(result) < min or float(result) > max:
                    
                    #     ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    #     ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((136,8,8))
                    # else :
                        ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                        ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((0,0,0))
                elif result == 'nogo' :
                    result='NG'
                    ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((136,8,8))
                
                else:
                    result='OK'
                    ws.cells(int(rstart),int(cstart)).value = str(result).upper()
                    ws.cells(int(rstart),int(cstart)).api.Font.Color=rgb_to_int((0,0,0))
                # ws.cells(int(rtime[0]),int(ctime[0])-1).value = all_mean
                # ws.cells(int(rstart[0]),int(cstart[0])+i).value = str(result[1])
                
                wb.save()
                t=False
    return check(request)

def wi(request):
    if request.method=='POST':
            model_sel=request.POST['model']
            try:
                order=[]
                data=md.objects.filter(model=model_sel).values('toolname').distinct()
                # cursor.execute("""SELECT DISTINCT Toolname From Model Where Model = ?   """,model_sel)
                tool = list(data)
            
            
                    
            
                tool_1=data[0]
                model=md.objects.distinct().values('model')
                # cursor.execute("""SELECT DISTINCT Model From Model """)
                # data = cursor.fetchall()
                # model=[data[0] for data in data[0:len(data)]]
            
                return render (request,'third version/wi.html',{'model':model,'tool':tool,'model_sel':model_sel,'order':tool})
            except:
                messages.info(request,'ไม่พบเครื่องมือของโมเดลนี้') 
    # cursor.execute("""SELECT DISTINCT Model From Model """) 
    # data = cursor.fetchall()
    # model=[data[0] for data in data[0:len(data)]]
    model=md.objects.distinct().values('model')
    return render (request,'third version/wi.html',{'model':model})

def wi_insert(request):
    wi_sel=request.POST.getlist('tool')
    model=request.POST['model']
    print(wi_sel)
   
    for count,value in enumerate(wi_sel,start=1):
        if value == '':
            continue
        elif wi_sel.count(value) >1:
            messages.info(request,'ผิดพลาด มีการกำหนดเครื่องมือซ้ำ')
            return wi(request)
        # cursor.execute(""" UPDATE Model SET wi = ? Where Toolname=? and Model = ?  """,(count,value,model))
        # cursor.commit()
        md.objects.filter(toolname=value,model=model).update(wi=count)
    messages.info(request,'Successful')
    return wi(request) 

def add_tool(request):
    if request.method=='POST':
        toolname=request.POST['toolname']
        series=request.POST['series']
        tooltype=request.POST['tooltype']
        tool=toolname+'_'+series
        if tooltype =='value':
            ttid=2
        else:ttid=1
        tl.objects.create(toolname=tool,ttid=ttid,tooltype=tooltype)
    #     cursor.execute('''INSERT INTO Tool ([Toolname]
    #   ,[ttid]
    #   ,[Tooltype]) VALUES (?,?,?)''',tool,ttid,tooltype)
    #     cursor.commit()
    return render (request,'third version/add_tool.html')

def csv(request):
    if request.method=='POST':
        try:
            uploaded=request.FILES['choose_file'] 
            fs=FileSystemStorage()
            fs.save(uploaded.name,uploaded)
            path='C:/excel_test/media/'+uploaded.name
            df=pd.read_csv(path,encoding = "ISO-8859-1")
            colunm=['Model'
                ,'Tooltype'
                ,'Cstart'
                ,'Rstart'
                ,'Excel'
                ,'img'
                ,'Point'
                ,'Toolname'
                ,'Machine_Line'
                ,'Machine_No'
                ,'Machine_Process'
                ,'Processname'
                ,'Partname'
                ,'Point_No'
                ,'Sheet'
                ,'shift'   
            ]
            df_data=df[colunm] 
            # recodes=[dict(zip(df_data.dtype.names,x)) for x  in df_data]
            # print(df_data.values)
            # df_data = df_data.fillna(value=0)
            recodes=df_data.values.tolist() 
            # print(recodes) 
            
            
#################################################################################################################
########################โค้ดการใส่ข้อมูลลงด้วยวิธีแบบ Model ใช้เวลา3วินาที#########################
            # list_data=[]
            # for v in recodes:
            #     dict={}
            #     for i, p in enumerate(colunm):
            #         p=p.lower()
            #         try:
            #                 dict[p]=v[i]
            #         except:
            #             continue
            #     list_data.append(dict)
          
            # for i in list_data:
            #     md.objects.create(**i,pattern='csv') 
            
####################################################################################################################

####################################################################################################################
########################โค้ดการใส่ข้อมูลลงด้วยวิธีแบบ Script ใช้เวลา0.5วินาที#########################
            sql='''INSERT INTO [S-pack].[dbo].[Model] ([Model]
        ,[Tooltype]
        ,[Cstart]
        ,[Rstart]
        ,[Excel]
        ,[img]
        ,[Point]
        ,[Toolname]
        ,[Machine_Line]
        ,[Machine_No]
        ,[Machine_Process]
        ,[Processname]
        ,[Partname]
        ,[Point_No]
        ,[Sheet]
        ,[shift],[pattern]) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
            
            cursor.fast_executemany = True
            for i in recodes:
                    dict={}
                    dict['model']=i[0]
                    dict['tooltype']=i[1]
                    
                    
                # try:
                    cursor.execute(sql,*i,'csv')
                    
                    cursor.commit()
 #################################################################################################       
            os.remove(path)  
            messages.info(request,'Successful')  
        except:
            messages.info(request,'ไฟล์มีข้อผิดพลาด') 
            os.remove(path)  
               
    # cursor.close()
    return render (request,'add/model_csv.html') 
 


# def editmodel(request):
    
#     if request.is_ajax(): 
#         print('axious')
#         partname = request.POST["partname"]
#         model =md.objects.filter(partname=partname).values('model').distinct()
#         model=list(model)
#         mo_list=[]
#         for i in model:
#             print(i)
#             mo_list.append(i.get('model'))
#         print(mo_list)
#         return JsonResponse( {'model':mo_list}) 
    
    
    
#     if request.method=='POST':
#         partname = request.POST["partname"]
#         model = request.POST["model"]
#         partname_list=md.objects.values('partname').distinct()
#         pattern =md.objects.filter(partname=partname,model=model).values('pattern').distinct()
#         point =md.objects.filter(partname=partname,model=model).values('point').distinct()
#         pattern_list=pt.objects.values('name').distinct()
#         print()
#         return render (request,'third version/edit_model.html',{'partname_sel':partname,'partname':partname_list,'model':model,'point':point,'pattern':pattern,'pattern_list':pattern_list})
        
#     partname =md.objects.values('partname').distinct()
  
#     return render (request,'third version/edit_model.html',{'partname':partname})

def overview(request):
    
    if request.is_ajax(): 
        print('axious')
        partname = request.POST["partname"]
        model =md.objects.filter(partname=partname).values('model').distinct()
        model=list(model)
        mo_list=[]
        for i in model:
      
            mo_list.append(i.get('model'))

        return JsonResponse( {'model':mo_list}) 
    
    
    
    if request.method=='POST':
        partname = request.POST["partname"]
        model = request.POST["model"]
        partname_list=md.objects.values('partname').distinct()
        pattern =md.objects.filter(partname=partname,model=model).values('pattern').distinct()
        point =md.objects.filter(partname=partname,model=model).values('point','toolname').distinct()
        pattern_list=pt.objects.values('name').distinct()
        print('part and model',partname,model)
        return render (request,'third version/overview.html',{'sel_partname':partname,'partname':partname_list,'model':model,'point':point,'pattern':pattern,'pattern_list':pattern_list})
        
    partname =md.objects.values('partname').distinct()
  
    return render (request,'third version/overview.html',{'partname':partname})
   
def pattern(request):
    if request.method=='POST': 
        try:
            test=request.POST['cstart']
            cstart=request.POST.getlist('cstart')
            rstart=request.POST.getlist('rstart')
            name_shift=request.POST.getlist('name_shift') 
            print(cstart,rstart,name_shift)
            name=request.POST['name']
            num_shift=request.POST['num_shift']
            direction=request.POST['direction']
            shift_type=request.POST.getlist('shift_type')
            timestart=request.POST.getlist('time_start')
            timeend=request.POST.getlist('time_end')
            print((timestart),(timeend))
            for i in range(int(num_shift)):
    
                pt.objects.create(name_shift=name_shift[i],cstart=column_index_from_string(cstart[i]),rstart=rstart[i],num_shift=num_shift,direction=direction,name=name,shift_type=shift_type[i],timestart=timestart[i],timeend=timeend[i])      
            
            print('hello')  
            # return render (request,'third version/pattern.html',{'name':name,'num_shift':num_shift,'direction':direction,'letter':letter})    
       
        except:
            name=request.POST['name']  
            num_shift=request.POST['num_shift']
            direction=request.POST['direction']  
            letter=[]
            print('kuy',num_shift,direction) 
            for i in range(int(num_shift)):
                letter.append(get_column_letter(i+1))
            print(letter)
            return render (request,'third version/pattern.html',{'name':name,'num_shift':num_shift,'direction':direction,'letter':letter})    
           
    return render (request,'third version/pattern.html') 

def delete(request):
    if request.is_ajax(): 
        print('axious')
        partname = request.POST["partname"]
        model =md.objects.filter(partname=partname).values('model').distinct()
        model=list(model)
        mo_list=[]
        for i in model:
      
            mo_list.append(i.get('model'))

        return JsonResponse( {'model':mo_list}) 
    
    
    
    if request.method=='POST':
        partname = request.POST["partname"]
        model = request.POST["model"]
    
        md.objects.filter(partname=partname,model=model).delete()
        wk.objects.filter(partname=partname,model=model,date=date.today()).delete()
        partname_list=md.objects.values('partname').distinct()
        messages.info(request,'Successful') 
        return render (request,'third version/delete.html',{'sel_partname':partname,'partname':partname_list,'model':model})
        
    partname =md.objects.values('partname').distinct()
  
    return render (request,'third version/delete.html',{'partname':partname})
    

# def change_pattern(request):
#     partname = request.POST["partname"]
#     model = request.POST["model"]
#     pattern=request.POST["pattern"] 
#     old_pattern=request.POST["old_pattern"]
    
#     direction=pt.objects.filter(name=old_pattern).values('direction').distinct()
 
#     print('di',direction[0]['direction'])
#     if direction[0]['direction'] == 'vertical':
#         try:
#             rstart_old=md.objects.filter(model=model,excel=excel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,pattern=sel_pattern,partname=partname).values('rstart').order_by('-rstart')[0:1]
#             rstart_new=rstart_old[0]
#             rstart_new=rstart_new['rstart']  
#             rstart_new=int(rstart_new)+1
#         except:
#             rstart_new= pt.objects.filter(name=sel_pattern).values('rstart')
#             rstart_new=int(rstart_new[0]['rstart'])

#         for i, p in enumerate(point):
#             shift=pt.objects.filter(name=sel_pattern).values('name_shift')
#             shift=list(shift)
#             print(shift)
#             if tooltype == 'value':
#                 for k in shift:
#                         cstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('cstart')
                  
#                         cstart=int(cstart[0]['cstart'])
#                         md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,cstart=cstart,rstart=rstart_new+i,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],min=min[i],max=max[i],partname=partname)
#                         # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',min=min[i],max=max[i],partname=partname)
#             else:
#                 for k in shift:
#                     print(sel_pattern,k['name_shift'])
#                     cstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('cstart')
#                     print(cstart)
#                     cstart=int(cstart[0]['cstart'])
#                     md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],partname=partname,cstart=cstart,rstart=rstart_new+i)
#                 # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',partname=partname)
#     elif direction[0] == 'horizontal':
#         try:
#             cstart_old=md.objects.filter(model=model,excel=excel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,pattern=sel_pattern,partname=partname).values('cstart').order_by('-cstart')[0:1]
        
#             cstart_new=cstart_old[0]
#             cstart_new=cstart_new['cstart']  
             
#             cstart_new=int(cstart_new)+1
#         except:
#             cstart_new= pt.objects.filter(name=sel_pattern).values('cstart')
#             cstart_new=int(cstart_new[0]['cstart'])
            
#         for i, p in enumerate(point):
#             shift=pt.objects.filter(name=sel_pattern).values('name_shift')
#             shift=list(shift)
#             print(shift)
#             if tooltype == 'value':
#                 for k in shift:
#                         rstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('rstart')
                  
#                         rstart=int(rstart[0]['rstart'])
#                         md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,rstart=rstart,cstart=cstart_new+i,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],min=min[i],max=max[i],partname=partname)
#                         # md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift='Night',min=min[i],max=max[i],partname=partname)
#             else:
#                 for k in shift:
#                     rstart=pt.objects.filter(name=sel_pattern,name_shift=k['name_shift']).values('rstart')       
#                     rstart=int(rstart[0]['rstart'])
#                     md.objects.create(model=model,excel=excel,tooltype=tooltype,pattern=sel_pattern,img=img,point=p,toolname=toolsel,machine_line=machname,machine_no=machno,machine_process =machpro,processname=process,sheet=sheet,shift=k['name_shift'],partname=partname,rstart=rstart,cstart=cstart_new+i)
 
#     md.objects.filter(pattern=old_pattern,partname=partname,model=model).update(pattern=pattern)
    
     
