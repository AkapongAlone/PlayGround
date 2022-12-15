
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, QObject, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QWidget, QPushButton, QApplication, QGridLayout, QMessageBox, QMainWindow)
from flask import Flask, render_template, url_for, request
import threading
import serial
import re
import time
import pyodbc
import datetime
from openpyxl import load_workbook, Workbook
import asyncio
import sys
import UI2 as main


wb = load_workbook('report.xlsx')
ws = wb.active
conn = pyodbc.connect(
    'driver=ODBC Driver 17 for SQL Server;server=DESKTOP-H45D3IV;database=measured_data;trusted_connection=yes;')
cursor = conn.cursor()

num_type = 7
amount = 8
i = 0
cursor.execute("""SELECT top (?) [min],[max],min(time)
    FROM [measured_data].[dbo].[Measured]
	
    GROUP BY min,max
	ORDER BY min(time)
	;""",  (num_type))
min_max = cursor.fetchall()
Vmin = [min_max[1] for min_max in min_max[0:len(min_max)]]
Vmax = [min_max[0] for min_max in min_max[0:len(min_max)]]
for i in range(0,len(Vmax)):
    ws.cell(row=12+2*i, column=(10)).value = "Max "+str(Vmax[i])
    ws.cell(row=13+2*i, column=(10)).value = "Min "+str(Vmin[i])

while i < num_type:
    cursor.execute("""SELECT Top (?) value FROM measured_data.dbo.Measured """,  (amount))
    Valueout = cursor.fetchall()
    
    allrow = [allrow[0] for allrow in Valueout[0:len(Valueout)]]
    
    for r in range(0, len(Valueout)):
        ws.cell(row=12+2*i, column=(11 + r)).value = allrow[r]
        
    
    i=i+1

    cursor.execute( """;WITH CTE AS(SELECT TOP (?) value FROM measured_data.dbo.Measured ) DELETE FROM CTE """,  (amount))  #CTEคือคำสั่งแนวExpressของsql

    
    
wb.save('report.xlsx')
