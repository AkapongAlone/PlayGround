
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, QObject, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QWidget, QPushButton, QApplication, QGridLayout, QMessageBox, QMainWindow)
from flask import Flask, render_template, url_for, request
from tkinter import *
from tkinter import simpledialog
import threading
import serial
import re
import time
import pyodbc
import datetime
from openpyxl import load_workbook, Workbook
import asyncio
import sys
import UI2 as main
import easygui


class Activity(QtWidgets.QMainWindow, main.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        # กำหนดค่าให้mainwindowsให้ใช้ฟังใช่พวกpyqtได้
        MainWindow = QtWidgets.QMainWindow()
        self.ui = main.Ui_MainWindow()  # ดึงคลาสจากUI2เข้ามาใช้ในคลาสนี้
        self.ui.setupUi(MainWindow)  # ผ่านค่าMainwindowsเข้าไปในคลาสsetupUi
        # เรียกใช้ฟังชั่นต่าง ในคลาสsetupUiเข้ามาใช้ในคลาสนี้ได้
        self.setupUi(self)
        self.bindWidget()
        self.mwidget = QMainWindow(self)
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        self.show()

        sys.exit(app.exec_())

    def bindWidget(self):

        self.process.clicked.connect(self.checkstatus)
        self.cancel.clicked.connect(self.end_program)
        self.Report.clicked.connect(self.create_report)

    def checkstatus(self):
        if self.amount.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกจำนวนชิ้นงาน", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.amount.setFocus()
        elif self.max.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกค่าmax", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.max.setFocus()
        elif self.min.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกค่าmin", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.min.setFocus()

        else:
            self.on_clickHaveThread()

    def end_program(self):
        sys.exit(app.exec_())

    def on_clickHaveThread(self):
        self.thread = BackgroundThread(
            self.min.text(), self.max.text(), self.amount.text())
        self.thread.start()
        self.thread.setTerminationEnabled(True)

        self.thread.result.connect(self.HaveThread)
        self.thread.countgo.connect(self.go_count)
        self.thread.countnogo.connect(self.nogo_count)
        self.thread.clear.connect(self.Clear)

        #self.thread.result.connect(self.updateProgressBar)

    def Clear(self, clear):
        buttonReply = QMessageBox.question(
            self, 'PyQt5 message', "กรุณากรอกข้อมูล", QMessageBox.Ok)
        self.min.setText(clear)
        self.max.setText(clear)
        self.amount.setText(clear)

    def HaveThread(self, result):
        self.value.setText(str(result))

    def go_count(self, countgo):
        self.Pass.setText(str(countgo))

    def nogo_count(self, countnogo):
        self.NotPass.setText(str(countnogo))

    def create_report(self):
        
        excelfile = Workbook()
        sheet = excelfile.active
        excelfile.save('requset.xlsx')
        self.export_xmls()

    def export_xmls(self):
        myvar = easygui.enterbox("What, is your favorite color?")

        wb = load_workbook('requset.xlsx')
        ws = wb.active
        conn = pyodbc.connect(
            'driver=ODBC Driver 17 for SQL Server;server=DESKTOP-H45D3IV;database=measured_data;trusted_connection=yes;')
        cursor = conn.cursor()

        cursor.execute("""SELECT value FROM measured_data.dbo.Measured """)

        Valueout = cursor.fetchall()
        print(len(Valueout))

        row1 = [row1[0] for row1 in Valueout[0:19]]  # แปลงpyodbc.Rowเป็นint
        row2 = [row2[0] for row2 in Valueout[19:38]]
        row3 = [row3[0] for row3 in Valueout[38:57]]
        row4 = [row4[0] for row4 in Valueout[57:76]]
        row5 = [row5[0] for row5 in Valueout[76:95]]
        row6 = [row6[0] for row6 in Valueout[95:114]]
        row8 = [row8[0] for row8 in Valueout[114:133]]
        row9 = [row9[0] for row9 in Valueout[133:152]]
        row10 = [row10[0] for row10 in Valueout[152:171]]
        row11 = [row11[0] for row11 in Valueout[171:190]]
        row12 = [row12[0] for row12 in Valueout[190:209]]

        cursor.execute("""SELECT go_nogo FROM measured_data.dbo.Measured """)
        result = cursor.fetchall()
        result1 = [result1[0] for result1 in result[0:19]]

        for i in range(0, len(row1)):
            ws.cell(row=12, column=(11 + i)).value = row1[i]

        for i in range(0, len(row2)):
            ws.cell(row=14, column=(11 + i)).value = row2[i]

        for i in range(0, len(row3)):
            ws.cell(row=16, column=(11 + i)).value = row3[i]

        for i in range(0, len(row4)):
            ws.cell(row=18, column=(11 + i)).value = row4[i]

        for i in range(0, len(row5)):
            ws.cell(row=20, column=(11 + i)).value = row5[i]

        for i in range(0, len(row6)):
            ws.cell(row=22, column=(11 + i)).value = row6[i]

        for i in range(0, len(result1)):
            ws.cell(row=24, column=(11 + i)).value = result1[i]

        for i in range(0, len(row8)):
            ws.cell(row=26, column=(11 + i)).value = row8[i]

        for i in range(0, len(row9)):
            ws.cell(row=28, column=(11 + i)).value = row9[i]

        for i in range(0, len(row10)):
            ws.cell(row=30, column=(11 + i)).value = row10[i]

        for i in range(0, len(row11)):
            ws.cell(row=32, column=(11 + i)).value = row11[i]

        for i in range(0, len(row12)):
            ws.cell(row=34, column=(11 + i)).value = row12[i]

        wb.save('requset.xlsx')
        QMessageBox.question(self, 'PyQt5 message',
                             "Successfull", QMessageBox.Ok)


class BackgroundThread(QtCore.QThread, ):

    result = QtCore.pyqtSignal(float)
    total = QtCore.pyqtSignal(int)
    countgo = QtCore.pyqtSignal(int)
    countnogo = QtCore.pyqtSignal(int)
    clear = QtCore.pyqtSignal(str)

    def __init__(self, limit_min, limit_max, amount):
        super().__init__()
        self.min = limit_min
        self.max = limit_max
        self.amount = amount

        self.threadactive = True

    def run(self):
        limit_min = self.min
        limit_max = self.max
        count = self.amount
        count = int(count)
        limit_min = float(limit_min)
        limit_max = float(limit_max)
        conn = pyodbc.connect(
            'driver=ODBC Driver 17 for SQL Server;server=DESKTOP-H45D3IV;database=measured_data;trusted_connection=yes;')

        cursor = conn.cursor()

        while 1:
            try:
                cursor.execute("""
                 CREATE TABLE Measured (
                    value       FLOAT ,
                    Tool        char(20),
                    time        DATETIME2,
                    go_nogo       VARCHAR (20),
                
                                    ) """)
                conn.commit()

            except:
                break

        tool = serial.Serial("COM4", 57600, timeout=10)
        lamp = serial.Serial("COM3")

        print(count)
        toolname_list = []
        go_list = []
        nogo_list = []

        i = 0

        while i < count:

            if tool.inWaiting():
                read_byte = tool.read(tool.inWaiting())

                print(read_byte)
                k = read_byte.decode("utf-8")
                print(k)
                k = k.split('+')
                if len(k) == 1:
                    continue
                if not k[0].startswith('D'):
                    print('error please try aging')
                    continue
                print(type(k[0]))
                toolname = k[0]

                try:

                    num = re.findall('[1-9]+.[0-9]', k[1])
                    value_list = [float(i) for i in num]

                    if value_list[0] > 100:
                        num = re.findall('[0]+.[0-9]', k[1])
                        value_list = [float(i) for i in num]

                    print(value_list[0])

                except:
                    print('try aging')
                    continue

                if value_list[0] >= limit_min and value_list[0] < limit_max:
                    go_list.append(value_list[0])
                    lamp.write('@??S001000!@??G!'.encode('ascii'))
                    print('go')
                    print('number go: ', len(go_list))
                    cursor.execute(
                        """SELECT * FROM measured_data.dbo.Measured """)
                    cursor.execute("""
                        INSERT INTO measured_data.dbo.Measured(value,Tool,time,go_nogo) 
                        VALUES (?,?,CURRENT_TIMESTAMP,?)""", (value_list[0], toolname, 'pass'))

                else:
                    nogo_list.append(value_list[0])
                    lamp.write('@??S100000!@??G!'.encode('ascii'))
                    print('no go')
                    print('number nogo: ', len(nogo_list))
                    cursor.execute(
                        """SELECT * FROM measured_data.dbo.Measured """)
                    cursor.execute("""
                        INSERT INTO measured_data.dbo.Measured(value,Tool,time,go_nogo) 
                        VALUES (?,?,CURRENT_TIMESTAMP,?)""", (value_list[0], toolname, 'fail'))

                conn.commit()
                time.sleep(0.5)
                lamp.write('@??S000000!@??G!'.encode('ascii'))
                print('total : '+str(len(go_list)+len(nogo_list)))
                self.result.emit(value_list[0])
                self.countgo.emit(len(go_list))
                self.countnogo.emit(len(nogo_list))

                i = i+1
        self.clear.emit(" ")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    maina = Activity()
    sys.exit(app.exec_())
