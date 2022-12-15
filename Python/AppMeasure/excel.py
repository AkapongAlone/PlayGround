import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.workbook import Workbook
import string
import getpass


class Excel(object):
    

    def creat(self,full_name):
        
        wb = Workbook()
        s = wb.active


        thin = Side(border_style="thin", color="000000")#Border style, color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)#Position of border


        for row in s['K11:AC41']:
        # ,'I36:I38','R3:R4']:
            for cell in row:
                cell.border = border#A5:D6 range cell setting border

        for row in s['J11:J38']:
        # ,'I36:I38','R3:R4']:
            for cell in row:
                cell.border = border#A5:D6 range cell setting border

        for row in s['I36:I38']:
            for cell in row:
                cell.border = border#A5:D6 range cell setting border

        for row in s['R3:R4']:
            for cell in row:
                cell.border = border#A5:D6 range cell setting border


        ##############################################################################


        BORDER_LIST = ['A1:B4','C1:T2','U1:W1','X1:Z1','AA1:AC1','C3:D4','E3:H4','I3:J3','I4:J4',
        'K3:Q3','K4:Q4','S3:T3','S3:T3','U2:W3','V4:W4','X2:Z3','Y4:Z4','AA2:AC3','AB4:AC4','A5:B5',
        'A6:B6','A7:B7','C5:H5','C6:H6','C7:H7','I5:J5','I6:J6','I7:J7','K5:Q5','K6:Q6','K7:Q7','R5:T5',
        'R6:T7','U5:W5','U6:W7','X5:Z5','X6:Z7','AA5:AC7','A9:B10','C9:E10','F9:J9','F10:J10','B11:H11',
        'A12:A13','B12:H12','B13:G13','I12:I13','A14:A15','B14:H14','B15:G15','A16:A17','B16:H16','B17:G17',
        'I16:I17','A18:A19','B19:G19','I18:I19','A20:A21','B20:H20','B21:G21','I20:I21','A22:A23','B22:H22',
        'B23:G23','I22:I23','A24:A25','B24:H24','B25:G25','I24:I25','A26:A27','B26:H26','B27:G27','I26:I27',
        'A28:A29','B28:H28','B29:G29','I28:I29','A30:A31','B30:H30','B31:G31','I30:I31','A32:A33','B32:H32',
        'B33:G33','I32:I33','A34:A35','B34:H34','B35:G35','I34:I35','B36:H36','B37:H37','B38:H38','B39:H39',
        'I39:J39','A40:J40','A41:J41','A44:D44','V44:AC44','I14:I15','S4:T4','K9:AC10','A1:AC44','A42:AC43',
        'A41:D43']





        def set_border(ws, cell_range):
            rows = ws[cell_range]
            side = Side(border_style='thin', color="FF000000")

            rows = list(rows)
            max_y = len(rows) - 1  # index of the last row
            for pos_y, cells in enumerate(rows):
                max_x = len(cells) - 1  # index of the last cell
                for pos_x, cell in enumerate(cells):
                    border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                    if pos_x == 0:
                        border.left = side
                    if pos_x == max_x:
                        border.right = side
                    if pos_y == 0:
                        border.top = side
                    if pos_y == max_y:
                        border.bottom = side

                    # set new border only if it's one of the edge cells
                    if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                        cell.border = border

        # border
        for pos in BORDER_LIST:
            set_border(s, pos)

        ##############################################################################


        s.merge_cells('A1:B4')

        s.merge_cells('C1:T2')
        cell1 = s.cell(row=1,column=3)
        cell1.value ='Working Form'
        cell1.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('U1:W1')
        cell2 = s.cell(row=1,column=21)
        cell2.value ='Approved by'
        cell2.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('X1:Z1')
        cell3 = s.cell(row=1,column=24)
        cell3.value ='Checked by'
        cell3.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('AA1:AC1')
        cell4 = s.cell(row=1,column=27)
        cell4.value ='Prepared'
        cell4.alignment = Alignment(vertical='center',horizontal='center')


        s.merge_cells('C3:D4')
        cell5 = s.cell(row=3,column=3)
        cell5.value ='Document name'
        cell5.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('E3:H4')
        cell6 = s.cell(row=3,column=5)
        cell6.value ='ใบตรวจสอบชิ้นงาน PCV 2W'
        cell6.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I3:J3')
        s.merge_cells('I4:J4')
        cell7 = s.cell(row=3,column=9)
        cell8 = s.cell(row=4,column=9)
        cell7.value ='Department'
        cell8.value ='Section'
        cell7.alignment = Alignment(vertical='center',horizontal='center')
        cell8.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('K3:Q3')
        s.merge_cells('K4:Q4')
        cell9 = s.cell(row=3,column=11)
        cell10 = s.cell(row=4,column=11)
        cell9.value ='Production'
        cell10.value ='NC'
        cell9.alignment = Alignment(vertical='center',horizontal='center')
        cell10.alignment = Alignment(vertical='center',horizontal='center')

        cell11 = s.cell(row=3,column=18)
        cell12 = s.cell(row=4,column=18)
        cell11.value ='Revision'
        cell12.value ='Page'
        cell11.alignment = Alignment(vertical='center',horizontal='center')
        cell12.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('S3:T3')
        s.merge_cells('S4:T4')
        cell13 = s.cell(row=3,column=19)
        cell14 = s.cell(row=4,column=19)
        cell13.value ='11'
        cell14.value ='1/1'
        cell13.alignment = Alignment(vertical='center',horizontal='center')
        cell14.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('U2:W3')
        s.merge_cells('V4:W4')
        cell13 = s.cell(row=3,column=21)
        cell14 = s.cell(row=4,column=21)
        cell14.value ='Date'
        cell13.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('X2:Z3')
        s.merge_cells('Y4:Z4')
        cell15 = s.cell(row=3,column=24)
        cell16 = s.cell(row=4,column=24)
        cell16.value ='Date'
        cell15.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('AA2:AC3')
        s.merge_cells('AB4:AC4')
        cell17 = s.cell(row=3,column=27)
        cell18 = s.cell(row=4,column=27)
        cell18.value ='Date'
        cell17.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('A5:B5')
        s.merge_cells('A6:B6')
        s.merge_cells('A7:B7')
        cell19 = s.cell(row=5,column=1)
        cell20 = s.cell(row=6,column=1)
        cell21 = s.cell(row=7,column=1)
        cell19.value ='Part Name'
        cell20.value ='Process'
        cell21.value ='Material'
        cell19.alignment = Alignment(vertical='center',horizontal='center')
        cell20.alignment = Alignment(vertical='center',horizontal='center')
        cell21.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('C5:H5')
        s.merge_cells('C6:H6')
        s.merge_cells('C7:H7')
        cell19 = s.cell(row=5,column=3)
        cell20 = s.cell(row=6,column=3)
        cell21 = s.cell(row=7,column=3)
        cell19.value ='PCV  2W   ( OP - 2 )'
        cell20.value ='NC Lathe'
        cell21.value ='SCM 415'
        cell19.alignment = Alignment(vertical='center',horizontal='center')
        cell20.alignment = Alignment(vertical='center',horizontal='center')
        cell21.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I5:J5')
        s.merge_cells('I6:J6')
        s.merge_cells('I7:J7')
        cell22 = s.cell(row=5,column=9)
        cell23 = s.cell(row=6,column=9)
        cell24 = s.cell(row=7,column=9)
        cell22.value ='Customer'
        cell23.value ='Customer Part No.'
        cell24.value ='TMT Part No.'
        cell22.alignment = Alignment(vertical='center',horizontal='center')
        cell23.alignment = Alignment(vertical='center',horizontal='center')
        cell24.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('K5:Q5')
        s.merge_cells('K6:Q6')
        s.merge_cells('K7:Q7')
        cell25 = s.cell(row=5,column=11)
        cell26 = s.cell(row=6,column=11)
        cell27 = s.cell(row=7,column=11)
        cell25.value ='SKD'
        cell26.value ='SM299224-0010'
        cell27.value ='SM299224-0010'
        cell25.alignment = Alignment(vertical='center',horizontal='center')
        cell26.alignment = Alignment(vertical='center',horizontal='center')
        cell27.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('R5:T5')
        s.merge_cells('R6:T7')
        s.merge_cells('U5:W5')
        s.merge_cells('U6:W7')
        s.merge_cells('X5:Z5')
        s.merge_cells('X6:Z7')
        s.merge_cells('AA5:AC7')
        cell28 = s.cell(row=5,column=18)
        cell29 = s.cell(row=5,column=21)
        cell30 = s.cell(row=5,column=24)
        cell31 = s.cell(row=5,column=27)
        cell28.value ='Reference :'
        cell29.value ='Lot number'
        cell30.value ='Machine No.'
        cell28.alignment = Alignment(vertical='center',horizontal='center')
        cell29.alignment = Alignment(vertical='center',horizontal='center')
        cell30.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('A9:B10')
        cell31 = s.cell(row=9,column=1)
        cell31.value ='Drawing No.'
        cell31.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('C9:E10')
        cell32 = s.cell(row=9,column=3)
        cell32.value ='SM299224-0010'
        cell32.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('F9:J9')

        s.merge_cells('F10:J10')
        cell33 = s.cell(row=10,column=6)
        cell33.value ='Date of production'
        cell33.alignment = Alignment(vertical='center',horizontal='center')

        cell34 = s.cell(row=11,column=1)
        cell34.value ='No'
        cell34.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B11:H11')
        cell35 = s.cell(row=11,column=2)
        cell35.value ='Description'
        cell35.alignment = Alignment(vertical='center',horizontal='center')

        cell36 = s.cell(row=11,column=9)
        cell36.value ='Tool'
        cell36.alignment = Alignment(vertical='center',horizontal='center')

        cell37 = s.cell(row=11,column=10)
        cell37.value ='Pcs / Check'
        cell37.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('A12:A13')
        cell38 = s.cell(row=12,column=1)
        cell38.value ='1'
        cell38.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B12:H12')
        cell39 = s.cell(row=12,column=2)
        cell39.value ='Length 1'

        s.merge_cells('B13:G13')
        cell40 = s.cell(row=13,column=2)
        cell40.value ='HG'

        cell41 = s.cell(row=13,column=8)
        cell41.value ='1/40'
        cell41.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I12:I13')
        cell42 = s.cell(row=12,column=9)
        cell42.value ='T6-Z,T8-Z'
        cell42.alignment = Alignment(vertical='center',horizontal='center')

        cell43 = s.cell(row=12,column=10)
        cell43.value ='Max 14.02'
        cell43.alignment = Alignment(vertical='center',horizontal='center')

        cell44 = s.cell(row=13,column=10)
        cell44.value ='Min  13.96'
        cell44.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A14:A15')
        cell45 = s.cell(row=14,column=1)
        cell45.value ='2'
        cell45.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B14:H14')
        cell46 = s.cell(row=14,column=2)
        cell46.value ='Length 2'

        s.merge_cells('B15:G15')
        cell47 = s.cell(row=15,column=2)
        cell47.value ='HG'

        cell48 = s.cell(row=15,column=8)
        cell48.value ='1/40'
        cell48.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I14:I15')
        cell49 = s.cell(row=14,column=9)
        cell49.value ='T1-Z'
        cell49.alignment = Alignment(vertical='center',horizontal='center')

        cell50 = s.cell(row=14,column=10)
        cell50.value ='Max 15.73'
        cell50.alignment = Alignment(vertical='center',horizontal='center')

        cell51 = s.cell(row=15,column=10)
        cell51.value ='Min  35.65'
        cell51.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A16:A17')
        cell52 = s.cell(row=16,column=1)
        cell52.value ='3'
        cell52.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B16:H16')
        cell53 = s.cell(row=16,column=2)
        cell53.value ='Length 3'

        s.merge_cells('B17:G17')
        cell54 = s.cell(row=17,column=2)
        cell54.value ='HG'

        cell55 = s.cell(row=17,column=8)
        cell55.value ='1/40'
        cell55.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I16:I17')
        cell56 = s.cell(row=16,column=9)
        cell56.value ='T1-Z'
        cell56.alignment = Alignment(vertical='center',horizontal='center')

        cell57 = s.cell(row=16,column=10)
        cell57.value ='Max 4.01'
        cell57.alignment = Alignment(vertical='center',horizontal='center')

        cell58 = s.cell(row=17,column=10)
        cell58.value ='Min  3.99'
        cell58.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A18:A19')
        cell59 = s.cell(row=18,column=1)
        cell59.value ='4'
        cell59.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B18:H18')
        cell60 = s.cell(row=18,column=2)
        cell60.value ='Outside Diameter 1'

        s.merge_cells('B19:G19')
        cell61 = s.cell(row=19,column=2)
        cell61.value ='MC'

        cell62 = s.cell(row=19,column=8)
        cell62.value ='1/40'
        cell62.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I18:I19')
        cell63 = s.cell(row=18,column=9)
        cell63.value ='T1-X'
        cell63.alignment = Alignment(vertical='center',horizontal='center')

        cell64 = s.cell(row=18,column=10)
        cell64.value ='Max 17.49'
        cell64.alignment = Alignment(vertical='center',horizontal='center')

        cell65 = s.cell(row=19,column=10)
        cell65.value ='Min  17.43'
        cell65.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A20:A21')
        cell66 = s.cell(row=20,column=1)
        cell66.value ='5'
        cell66.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B20:H20')
        cell67 = s.cell(row=20,column=2)
        cell67.value ='Outside Diameter 3'

        s.merge_cells('B21:G21')
        cell68 = s.cell(row=21,column=2)
        cell68.value ='MC'

        cell69 = s.cell(row=21,column=8)
        cell69.value ='1/40'
        cell69.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I20:I21')
        cell70 = s.cell(row=20,column=9)
        cell70.value ='T2-X'
        cell70.alignment = Alignment(vertical='center',horizontal='center')

        cell71 = s.cell(row=20,column=10)
        cell71.value ='Max 19.89'
        cell71.alignment = Alignment(vertical='center',horizontal='center')

        cell72 = s.cell(row=21,column=10)
        cell72.value ='Min  19.87'
        cell72.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A22:A23')
        cell73 = s.cell(row=22,column=1)
        cell73.value ='6'
        cell73.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B22:H22')
        cell74 = s.cell(row=22,column=2)
        cell74.value ='Inside Diameter2'

        s.merge_cells('B23:G23')
        cell75 = s.cell(row=23,column=2)
        cell75.value ='Bore Gauge'

        cell76 = s.cell(row=23,column=8)
        cell76.value ='1/40'
        cell76.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I22:I23')
        cell77 = s.cell(row=22,column=9)
        cell77.value ='T8-X'
        cell77.alignment = Alignment(vertical='center',horizontal='center')

        cell78 = s.cell(row=22,column=10)
        cell78.value ='Max 16.31'
        cell78.alignment = Alignment(vertical='center',horizontal='center')

        cell79 = s.cell(row=23,column=10)
        cell79.value ='Min  26,28'
        cell79.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A24:A25')
        cell80 = s.cell(row=24,column=1)
        cell80.value ='7'
        cell80.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B24:H24')
        cell81 = s.cell(row=24,column=2)
        cell81.value ='Inside Diameter 3.41'

        s.merge_cells('B25:G25')
        cell82 = s.cell(row=25,column=2)
        cell82.value ='เช็ค พินเกจ'

        cell83 = s.cell(row=25,column=8)
        cell83.value ='1/40'
        cell83.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I24:I25')
        cell84 = s.cell(row=24,column=9)
        cell84.value ='T7-Z'
        cell84.alignment = Alignment(vertical='center',horizontal='center')

        cell85 = s.cell(row=24,column=10)
        cell85.value ='Max 16.31'
        cell85.alignment = Alignment(vertical='center',horizontal='center')

        cell86 = s.cell(row=25,column=10)
        cell86.value ='Min  26,28'
        cell86.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A26:A27')
        cell87 = s.cell(row=26,column=1)
        cell87.value ='8'
        cell87.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B26:H26')
        cell88 = s.cell(row=26,column=2)
        cell88.value ='Length'

        s.merge_cells('B27:G27')
        cell89 = s.cell(row=27,column=2)
        cell89.value ='HG'

        cell90 = s.cell(row=27,column=8)
        cell90.value ='1/40'
        cell90.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I26:I27')
        cell91 = s.cell(row=26,column=9)
        cell91.value ='T8-X'
        cell91.alignment = Alignment(vertical='center',horizontal='center')

        cell92 = s.cell(row=26,column=10)
        cell92.value ='Max 4.53'
        cell92.alignment = Alignment(vertical='center',horizontal='center')

        cell93 = s.cell(row=27,column=10)
        cell93.value ='Min 4.47'
        cell93.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A28:A29')
        cell94 = s.cell(row=28,column=1)
        cell94.value ='9'
        cell94.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B28:H28')
        cell95 = s.cell(row=28,column=2)
        cell95.value ='ChamFer'

        s.merge_cells('B29:G29')
        cell96 = s.cell(row=29,column=2)
        cell96.value ='2D'

        cell97 = s.cell(row=29,column=8)
        cell97.value ='3/SHIFT'
        cell97.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I28:I29')
        cell98 = s.cell(row=28,column=9)
        cell98.value ='T8-X'
        cell98.alignment = Alignment(vertical='center',horizontal='center')

        cell99 = s.cell(row=28,column=10)
        cell99.value ='Max 13.93'
        cell99.alignment = Alignment(vertical='center',horizontal='center')

        cell100 = s.cell(row=29,column=10)
        cell100.value ='Min  13.87'
        cell100.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A30:A31')
        cell101 = s.cell(row=30,column=1)
        cell101.value ='10'
        cell101.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B30:H30')
        cell102 = s.cell(row=30,column=2)
        cell102.value ='ChamFer'

        s.merge_cells('B31:G31')
        cell103 = s.cell(row=31,column=2)
        cell103.value ='2D'

        cell104 = s.cell(row=31,column=8)
        cell104.value ='3/SHIFT'
        cell104.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I30:I31')
        cell105 = s.cell(row=30,column=9)
        cell105.value ='T8-X'
        cell105.alignment = Alignment(vertical='center',horizontal='center')

        cell106 = s.cell(row=30,column=10)
        cell106.value ='Max 12.63'
        cell106.alignment = Alignment(vertical='center',horizontal='center')

        cell107 = s.cell(row=31,column=10)
        cell107.value ='Min  12.57'
        cell107.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A32:A33')
        cell108 = s.cell(row=32,column=1)
        cell108.value ='11'
        cell108.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B32:H32')
        cell109 = s.cell(row=32,column=2)
        cell109.value ='ChamFer'

        s.merge_cells('B33:G33')
        cell110 = s.cell(row=33,column=2)
        cell110.value ='2D'

        cell111 = s.cell(row=33,column=8)
        cell111.value ='3/SHIFT'
        cell111.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I32:I33')
        cell119 = s.cell(row=32,column=9)
        cell119.value ='T8-X'
        cell119.alignment = Alignment(vertical='center',horizontal='center')

        cell113 = s.cell(row=32,column=10)
        cell113.value ='Max 3.90'
        cell113.alignment = Alignment(vertical='center',horizontal='center')

        cell114 = s.cell(row=33,column=10)
        cell114.value ='Min 3.70'
        cell114.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A34:A35')
        cell115 = s.cell(row=34,column=1)
        cell115.value ='12'
        cell115.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B34:H34')
        cell116 = s.cell(row=34,column=2)
        cell116.value ='Inside Diameter'

        s.merge_cells('B35:G35')
        cell117 = s.cell(row=35,column=2)
        cell117.value ='2D'

        cell118 = s.cell(row=35,column=8)
        cell118.value ='3/SHIFT'
        cell118.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('I34:I35')
        cell119 = s.cell(row=34,column=9)
        cell119.value ='T7-X'
        cell119.alignment = Alignment(vertical='center',horizontal='center')

        cell120 = s.cell(row=34,column=10)
        cell120.value ='Max 3.45'
        cell120.alignment = Alignment(vertical='center',horizontal='center')

        cell121 = s.cell(row=35,column=10)
        cell121.value ='Min 3.37'
        cell121.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        cell122 = s.cell(row=36,column=1)
        cell122.value ='13'
        cell122.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B36:H36')
        cell123 = s.cell(row=36,column=2)
        cell123.value ='ผิวชิ้นงาน  Burr,Scratch,Peeling'

        cell124 = s.cell(row=36,column=9)
        cell124.value ='1/40'
        cell124.alignment = Alignment(vertical='center',horizontal='center')

        cell125 = s.cell(row=36,column=10)
        cell125.value ='Visual'
        cell125.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        cell126 = s.cell(row=37,column=1)
        cell126.value ='14'
        cell126.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B37:H37')
        cell127 = s.cell(row=37,column=2)
        cell127.value ='ผิวชิ้นงาน  Burr,Scratch,Peeling'

        cell128 = s.cell(row=37,column=9)
        cell128.value ='1/40'
        cell128.alignment = Alignment(vertical='center',horizontal='center')

        cell129 = s.cell(row=37,column=10)
        cell129.value ='Visual'
        cell129.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        cell130 = s.cell(row=38,column=1)
        cell130.value ='15'
        cell130.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B38:H38')
        cell131 = s.cell(row=38,column=2)
        cell131.value ='ตรวจเช็คเกลียว'

        cell132 = s.cell(row=38,column=9)
        cell132.value ='1/40'
        cell132.alignment = Alignment(vertical='center',horizontal='center')

        cell133 = s.cell(row=38,column=10)
        cell133.value ='Visual'
        cell133.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        cell134 = s.cell(row=39,column=1)
        cell134.value ='16'
        cell134.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('B39:H39')
        cell135 = s.cell(row=39,column=2)
        cell135.value ='ตรวจเช็คท่อลม, ท่อน้ำมัน ทุกครั้งหลังจากเปลี่ยนมีด'

        s.merge_cells('I39:J39')
        cell136 = s.cell(row=39,column=9)
        cell136.value ='Visual'
        cell136.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A40:J40')
        cell137 = s.cell(row=40,column=1)
        cell137.value ='Check by'
        cell137.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('A41:J41')
        cell138 = s.cell(row=41,column=1)
        cell138.value ='RECHECK BY (ในกรณีที่เปลี่ยนมีดกลึง)'
        cell138.alignment = Alignment(vertical='center',horizontal='center')

        ##############################################################################

        s.merge_cells('A42:D42')
        cell139 = s.cell(row=42,column=1)
        cell139.value ='ชื่อเครื่องมือวัด'
        cell139.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('E42:AC42')
        cell140 = s.cell(row=42,column=5)
        cell140.value ='*ให้พนักงานทำการวัดค่าชิ้นงานทุกจุดก่อนและหลังการเปลี่ยนมีดและหัวหน้างานต้องทำการตรวจสอบทุกครั้ง'
        cell140.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('E43:AC43')
        cell141 = s.cell(row=43,column=5)
        cell141.value ='**พนักงานต้องทำการวัดชิ้นงานชิ้นแรก ทุกๆ40ชิ้น และชิ้นสุดท้ายที่ทำการผลิต'
        cell141.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('A44:D44')
        cell142 = s.cell(row=44,column=1)
        cell142.value ='FM-NC-022'
        cell142.alignment = Alignment(vertical='center',horizontal='center')

        s.merge_cells('V44:AC44')
        cell143 = s.cell(row=44,column=22)
        cell143.value ='Rev. 11  Effective Date: 5/2/2562'
        cell143.alignment = Alignment(horizontal='right')

        wb.save('C:\\Users\\user\\Desktop\\save file\\'+ full_name)





















