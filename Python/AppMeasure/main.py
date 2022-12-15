
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, QObject, pyqtSignal, pyqtSlot,Qt
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (QWidget, QPushButton, QApplication, QGridLayout, QMessageBox, QMainWindow)

from tkinter import *
from tkinter import simpledialog
import threading
import serial
import re
import time
import pyodbc
import datetime
from openpyxl import load_workbook,Workbook
import asyncio
import sys 
import UI4 as main
import easygui
import excel as report
import os
from datetime import date


class Activity(QtWidgets.QMainWindow, main.cssden, report.Excel):
    def __init__(self):
        super().__init__()
        MainWindow = QtWidgets.QMainWindow()     #กำหนดค่าให้mainwindowsให้ใช้ฟังใช่พวกpyqtได้
        self.ui = main.cssden()  # ดึงคลาสจากUI2เข้ามาใช้ในคลาสนี้
        self.ui.setupUi(MainWindow)              #ผ่านค่าMainwindowsเข้าไปในคลาสsetupUi
        self.setupUi(self)                       #เรียกใช้ฟังชั่นต่าง ในคลาสsetupUiเข้ามาใช้ในคลาสนี้ได้
        self.excel = report.Excel()
        self.bindWidget()
        self.mwidget = QMainWindow(self)
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        def moveWindow(event):
                        # RESTORE BEFORE MOVE
                        if  self.returnStatus() == 1:
                            self.minizie_or_maximize_window(self)

                        # IF LEFT CLICK MOVE WINDOW
                        if event.buttons() == Qt.LeftButton:
                            self.move(self.pos() + event.globalPos() - self.dragPos)
                            self.dragPos = event.globalPos()
                            event.accept()

        self.title_bar.mouseMoveEvent = moveWindow
        self.show()
        
        
        
        sys.exit(app.exec_())
        
    def bindWidget(self):

        
        self.process.clicked.connect(self.checkstatus)
        self.cancel.clicked.connect(self.end_program)
        self.Report.clicked.connect(self.create_report)
        
    def checkstatus(self):
        if  self.amount.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกจำนวนชิ้นงาน", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.amount.setFocus()
        elif self.max.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกค่าmax", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.max.setFocus()
        elif self.min.text() == '':
            QMessageBox.question(self, 'PyQt5 message',
                                 "กรอกค่าmin", QMessageBox.Ok)
            print("Empty Value Not Allowed")
            self.min.setFocus()

        else:
            self.on_clickHaveThread()
    
      

    def end_program(self):
        sys.exit(app.exec_())

    def bug(self,error):
        self.checkstatus()

    def on_clickHaveThread(self):
        self.thread = BackgroundThread(self.min.text(),self.max.text(),self.amount.text())
        self.thread.start()
        self.thread.setTerminationEnabled(True)

        self.thread.result.connect(self.HaveThread)
        self.thread.countgo.connect(self.go_count)
        self.thread.countnogo.connect(self.nogo_count)
        self.thread.clear.connect(self.Clear)
        self.thread.error.connect(self.bug)
        self.thread.totall.connect(self.total1)
        
        #self.thread.result.connect(self.updateProgressBar)

    def Clear(self,clear):
        self.killthread()
        QMessageBox.question(self, 'PyQt5 message', "กรุณากรอกข้อมูล", QMessageBox.Ok )
        self.min.setText(clear)
        self.max.setText(clear)
        self.amount.setText(clear)
        

    def killthread(self):
        self.thread.stop()
        
        print('How do I do this')
    
    def HaveThread(self, result):
        self.value.setText(str(result))

    def go_count(self,countgo): 
        self.Pass.setText(str(countgo))

    def nogo_count(self,countnogo):
        self.NotPass.setText(str(countnogo))

    def total1(self,totall):
        self.total.setText(str(totall))

    def create_report(self):
        full_name = easygui.enterbox("ใส่ชื่อ")+'_'+str(date.today())+".xlsx"
        self.excel.creat(full_name)
        # excelfile=Workbook()
        # sheet=excelfile.active
        # excelfile.save('report.xlsx')
        self.export_xmls(full_name)
              
    def export_xmls(self, full_name):
        
        path = 'C:\\Users\\user\\Desktop\\save file\\' + full_name
        num_type = easygui.enterbox("ใส่จำนวนชนิดชิ้นงาน")
        amount = easygui.enterbox("ใส่จำนวนชิ้นงาน")
        while True:            
            try:
                num_type = int(num_type)
                amount = int(amount)
                              
            except:
                QMessageBox.question(self, 'PyQt5 message', "กรุณากรอกข้อมูลตัวเลข", QMessageBox.Ok)
                break
            
            wb = load_workbook(os.path.join(path))
            ws = wb.active
            conn = pyodbc.connect(
                "driver=ODBC Driver 17 for SQL Server;server=DESKTOP-H45D3IV;database=measured_data;trusted_connection=yes;")
            cursor = conn.cursor()
            cursor.execute("""SELECT top (?) [min],[max],min(time)
            FROM [measured_data].[dbo].[Measured]
            
            GROUP BY min,max
            ORDER BY min(time)
            ;""",  (num_type))


            min_max = cursor.fetchall()
            Vmin = [min_max[1] for min_max in min_max[0:len(min_max)]]
            Vmax = [min_max[0] for min_max in min_max[0:len(min_max)]]
            i = 0
            for t in range(0, len(Vmax)):
                ws.cell(row=12+2*t, column=(10)).value = "Max "+str(Vmax[t])
                ws.cell(row=13+2*t, column=(10)).value = "Min "+str(Vmin[t])
            while i < num_type:
            
                cursor.execute(
                    """SELECT Top (?) value FROM measured_data.dbo.Measured """,  (amount))
                Valueout = cursor.fetchall()

                allrow = [allrow[0] for allrow in Valueout[0:len(Valueout)]]
               
                for r in range(0, len(Valueout)):
                    ws.cell(row=12+2*i, column=(11 + r)).value = allrow[r]
                   

                i = i+1
                cursor.execute(""";WITH CTE AS(SELECT TOP (?) value FROM measured_data.dbo.Measured ) DELETE FROM CTE """,
                            (amount))  # CTEคือคำสั่งแนวExpressของsql

            wb.save(path)
            QMessageBox.question(self, 'PyQt5 message',
                                "successful", QMessageBox.Ok)
            return False


class BackgroundThread(QtCore.QThread, ):
    
    result = QtCore.pyqtSignal(float)
    totall = QtCore.pyqtSignal(int)
    countgo = QtCore.pyqtSignal(int)
    countnogo = QtCore.pyqtSignal(int)
    clear = QtCore.pyqtSignal(str)
    error = QtCore.pyqtSignal(str)

    def __init__(self, limit_min, limit_max, amount):
        super().__init__( )
        self.min = limit_min
        self.max = limit_max
        self.amount = amount
        self.check= 5
        self.threadactive = True

    def run(self):
        
        i=0
        while self.threadactive:
            try:
                limit_min = self.min
                limit_max = self.max
                count = self.amount
                count = int(count)
                limit_min = float(limit_min)
                limit_max = float(limit_max)
            except:
                self.threadactive = FALSE
                
                break
         

            conn = pyodbc.connect(
                        "driver=ODBC Driver 17 for SQL Server;server=DESKTOP-H45D3IV;database=measured_data;trusted_connection=yes;")

            cursor = conn.cursor()
                    
            while 1:
                try:
                    cursor.execute("""
                        CREATE TABLE Measured (
                            value       FLOAT ,
                            min         FLOAT,
                            max         FLOAT,
                            Tool        char(20),
                            time        DATETIME,
                            Date        DATE,
                            go_nogo       VARCHAR (20),
                            
                        
                                            ) """)
                    conn.commit()   
                                
                except:
                    break
                    
            tool = serial.Serial("COM4", 57600, timeout=10)
            lamp = serial.Serial("COM3")
                    
                
            toolname_list = []
            go_list = []
            nogo_list = []
            
            
                
            while i < count:
                
                if limit_min >= limit_max:
                    self.clear.emit(" ")
                    break
        
                if tool.inWaiting():
                    read_byte = tool.read(tool.inWaiting())

                    print(read_byte)
                    k = read_byte.decode("utf-8")
                    print(k)
                    k = k.split('+')
                    if len(k) == 1:
                        continue
                    if not k[0].startswith('D'):
                        print('error please try aging')
                        continue
                    print(type(k[0]))
                    toolname = k[0]

                    try:

                        num = re.findall('[1-9]+.[0-9]+', k[1])
                        value_list = [float(i) for i in num]

                        if value_list[0] > 100:
                            num = re.findall('[0]+.[0-9]+', k[1])
                            value_list = [float(i) for i in num]

                        print(value_list[0])


                    except:
                        print('try aging')
                        continue

                    if value_list[0] >= limit_min and value_list[0] < limit_max:
                        go_list.append(value_list[0])
                        lamp.write('@??S001000!@??G!'.encode('ascii'))
                        print('go')
                        print('number go: ', len(go_list))
                        cursor.execute(
                                    """SELECT * FROM measured_data.dbo.Measured """)
                        cursor.execute("""
                                INSERT INTO measured_data.dbo.Measured(value,min,max,Tool,time,Date,go_nogo) 
                                VALUES (?,?,?,?,GETDATE(),CAST(GETDATE() AS Date),?)""", (value_list[0], limit_min, limit_max, toolname, 'pass'))

                                

                    else:
                        nogo_list.append(value_list[0])
                        lamp.write('@??S100000!@??G!'.encode('ascii'))
                        print('no go')
                        print('number nogo: ', len(nogo_list))
                        cursor.execute(
                                    """SELECT * FROM measured_data.dbo.Measured """)
                        cursor.execute("""
                                INSERT INTO measured_data.dbo.Measured(value,min,max,Tool,time,Date,go_nogo) 
                                VALUES (?,?,?,?,GETDATE(),CAST(GETDATE() AS Date),?)""", (value_list[0], limit_min, limit_max, toolname, 'fail'))

                    conn.commit()
                    time.sleep(0.5)
                    lamp.write('@??S000000!@??G!'.encode('ascii'))
                    print('total : '+str(len(go_list)+len(nogo_list)))
                    self.totall.emit(len(go_list)+len(nogo_list))
                    self.result.emit(value_list[0])
                    self.countgo.emit(len(go_list))
                    self.countnogo.emit(len(nogo_list))
                            
                            
                    i = i+1
            tool.close()
            lamp.close()
            self.threadactive=FALSE
            self.clear.emit(" ")
        
    def stop(self):
        print('eiei')
        self.threadactive = False
        self.wait()


        

     


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    maina = Activity()
    sys.exit(app.exec_())
